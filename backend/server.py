from fastapi import FastAPI, APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import csv
import json
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Literal
import uuid
from datetime import datetime, timezone
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Phase 6 · Slice 2 — dashboard endpoints now derive all payment / mode /
# active-record filtering through the shared domain layer. See
# /app/memory/phase6_shared_domain_preimpl_report.md.
from domain import (
    from_paise, to_paise,
    is_customer_payment_active, is_purchase_payment_active,
    is_cash_book_entry_canonical,
    sum_received_kpi, sum_paid_kpi, sum_mode_totals,
    sum_allocations_to_order, sum_allocations_to_purchase,
    compute_party_metrics,
    order_realized_amounts, order_estimated_amounts, order_unrealized,
    order_outstanding_from_alloc,
    purchase_realized_amounts, purchase_outstanding_from_alloc,
    _order_shipped_qty_by_item,
    # Bug fix (2026-07-22) — see domain.py for full context
    order_dashboard_outstanding_paise,
    sum_dashboard_outstanding_receivable_paise,
    derive_completion_shipped_date,
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="Artisan Ledger — Order-based P&L")
api_router = APIRouter(prefix="/api")

# ─── Local module imports (defined here so all endpoints can use them) ─────
from party_ledger_v2 import make_router as make_party_ledger_v2_router, ensure_bootstrap as ensure_party_ledger_bootstrap  # noqa: E402
from party_sync import (                                                                                                     # noqa: E402
    get_or_create_customer_party, get_or_create_vendor_party,
    sync_vendor_directory, rename_party, resolve_party,
    run_party_migration, is_ff_alias, SYSTEM_FF_ID,
)
from transfers import (                                                                                                       # noqa: E402
    Transfer, TransferIn, TransferSide,
    create_transfer, edit_transfer, reverse_transfer,
    derive_account_balance, ff_settlement_delta_from_transfers,
    run_transfer_migration, ensure_transfer_indexes,
)
from auth import (                                                                                                            # noqa: E402
    hash_password, verify_password, create_access_token, create_refresh_token,
    decode_token, new_user_doc, user_public, set_auth_cookies, clear_auth_cookies,
    get_current_user_from_db, require_admin,
)
from admin_reset import (                                                                                                     # noqa: E402
    preview_reset, execute_reset, create_backup, list_backups, get_backup_meta,
    delete_backup, load_test_dataset, remove_test_dataset,
    list_audit_logs, log_audit,
    is_reset_enabled, current_environment,
    BACKUP_DIR, APP_VERSION,
)
from reconcile import run_reconcile, summarize as summarize_reconcile           # noqa: E402


# ================================================================
# MODELS
# ================================================================
FACTORY_SUPPLIER_ID = "factory"
FACTORY_PARTY_NAME = "Father's Firm"


class PurchaseSource(BaseModel):
    """One purchase source row inside an OrderItem — a single supplier that
    contributed some or all of the item's Complete / Glass / Fitting cost.

    supplier_id == FACTORY_SUPPLIER_ID means the item's Factory (Father's Firm)
    provided that portion — the backend maps it to the canonical Father's Firm
    party for settlement purposes.
    """
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    supplier_id: str = ""
    supplier_name: str = ""
    complete: float = 0
    glass: float = 0
    fitting: float = 0


class OrderItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    main_category: str
    sub_category: Optional[str] = ""
    product_name: str
    model: Optional[str] = ""  # future: model/design
    qty: float = 0
    rate: float = 0
    product_sales: float = 0
    # Unified per-supplier purchase rows. When present + non-empty, the legacy
    # factory_*/outside_* fields are recomputed from these on every save so
    # dashboard KPIs and cost aggregates stay in lock-step with the new UI.
    purchase_sources: List[PurchaseSource] = []
    # Legacy split — retained as denormalised sums for aggregate/dashboard code.
    factory_complete: float = 0
    factory_glass: float = 0
    factory_fitting: float = 0
    outside_complete: float = 0
    outside_glass: float = 0
    outside_fitting: float = 0


class AdjustmentEntry(BaseModel):
    """Generic revenue or expense adjustment. Future-ready — attach `kind`
    later (e.g. discount, tip, loose_sale) without changing schema."""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    description: str
    amount: float = 0


PaymentStatus = Literal["Paid", "Partial", "Unpaid"]
TaxType = Literal["None", "GST", "IGST", "CGST_SGST"]

PAYMENT_MODES = ["Cash", "UPI", "Bank Transfer", "Cheque", "Credit Card", "Debit Card", "Other"]
ACCOUNT_TYPES = ["Bank", "Cash", "PettyCash", "UPI", "Wallet", "Gateway", "Other"]
ORDER_STATUSES = ["Draft", "Confirmed", "Packed", "Partially Shipped", "Fully Shipped", "Delivered", "Cancelled"]


class ShipmentItem(BaseModel):
    """A line item inside a Shipment — references an Order item by id."""
    model_config = ConfigDict(extra="ignore")
    order_item_id: str
    qty: float = 0


class Shipment(BaseModel):
    """A single dispatch event against an Order. Multiple allowed per order."""
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: Optional[str] = None
    items: List[ShipmentItem] = []
    boxes_shipped: float = 0
    freight_charged: float = 0
    freight_paid: float = 0
    transporter: Optional[str] = ""
    lr_number: Optional[str] = ""
    remarks: Optional[str] = ""


class Account(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: str = "Bank"
    notes: Optional[str] = ""
    archived: bool = False
    # Phase 3: optional opening balance (used by derive_account_balance).
    opening_balance: float = 0.0
    opening_date: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class PaymentAllocation(BaseModel):
    """One-payment-to-many-orders allocation. Stored embedded on CustomerPayment."""
    model_config = ConfigDict(extra="ignore")
    order_id: str
    amount: float = 0


class CustomerPaymentBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    customer_name: str
    date: Optional[str] = None
    amount: float = 0
    mode: str = "UPI"
    account_id: Optional[str] = ""
    account_name: Optional[str] = ""
    reference: Optional[str] = ""
    remarks: Optional[str] = ""
    allocations: List[PaymentAllocation] = []
    # Party-ledger linkage — when the money was received by a third party
    # (e.g. Father's Firm), the party ledger will auto-post a matching
    # linked effect against that party's settlement.
    received_by_party_id: Optional[str] = None
    received_by_party_name: Optional[str] = None


class CustomerPayment(CustomerPaymentBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    # computed
    allocated_total: float = 0
    unallocated: float = 0  # advance
    # Phase 2: stable party id (source-of-truth identity)
    customer_party_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class OrderBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    client_name: str
    order_date: Optional[str] = None
    shipped_date: Optional[str] = None  # deprecated (now derived from shipments) but kept
    status: str = "Confirmed"  # new: Draft/Confirmed/Packed/Partially Shipped/Fully Shipped/Delivered/Cancelled
    payment_status: PaymentStatus = "Unpaid"  # derived from customer_payments allocations
    notes: Optional[str] = ""

    items: List[OrderItem] = []
    shipments: List[Shipment] = []

    # Packing (order-level — same as before)
    boxes_used: float = 0
    cost_per_box: float = 0
    packing_cost: float = 0

    # Freight — legacy order-level values, kept for backward-compat but shipments own the source of truth
    boxes_shipped: float = 0
    freight_charged: float = 0
    freight_paid: float = 0
    transporter: Optional[str] = ""
    lr_number: Optional[str] = ""

    # Future-ready adjustments — generic descriptor+amount rows
    other_revenue: List[AdjustmentEntry] = []
    other_expense: List[AdjustmentEntry] = []

    # Legacy revenue placeholders (kept for backwards-compat)
    packing_recovery: float = 0
    other_charges: float = 0

    # Tax
    tax_applicable: bool = False
    tax_type: TaxType = "None"
    tax_percent: float = 0
    tax_amount: float = 0
    tax_amount_manual: bool = False


class Order(OrderBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    # Phase 2: stable party id (source-of-truth identity)
    customer_party_id: Optional[str] = None
    # ORDERED aggregates (potential — informational)
    ordered_qty_total: float = 0
    ordered_product_sales: float = 0

    # SHIPPED aggregates (recognized — used everywhere for revenue/profit)
    shipped_qty_total: float = 0
    shipped_product_sales: float = 0
    product_sales_total: float = 0     # alias of shipped_product_sales (legacy name used by dashboard/UI)
    factory_cost_total: float = 0
    outside_cost_total: float = 0
    other_revenue_total: float = 0
    other_expense_total: float = 0
    ship_freight_charged_total: float = 0
    ship_freight_paid_total: float = 0
    ship_boxes_shipped_total: float = 0
    operating_revenue: float = 0
    total_cost: float = 0
    invoice_total: float = 0
    net_profit: float = 0
    margin_percent: float = 0
    shipment_progress_percent: float = 0  # shipped_qty / ordered_qty * 100

    # Phase 4 — ESTIMATED (potential) aggregates: what the order will yield
    # if every remaining unshipped unit is eventually shipped at the
    # committed rates. `realized_*` are the same numbers already exposed above
    # (operating_revenue / net_profit) — kept as explicit aliases for the UI.
    estimated_factory_cost_total: float = 0
    estimated_outside_cost_total: float = 0
    estimated_operating_revenue: float = 0
    estimated_total_cost: float = 0
    estimated_net_profit: float = 0
    estimated_margin_percent: float = 0
    realized_revenue: float = 0          # alias of operating_revenue
    realized_net_profit: float = 0       # alias of net_profit
    revenue_recognized: float = 0        # PRD-mandated name = operating_revenue
    unrealized_revenue: float = 0        # estimated_operating_revenue - operating_revenue
    unrealized_net_profit: float = 0     # estimated_net_profit - net_profit

    # Payment aggregates (from customer_payments allocations)
    total_received: float = 0
    outstanding_balance: float = 0

    # Auto-derived latest shipped date (for month bucketing)
    last_shipped_date: Optional[str] = None

    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class PaymentBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    date: Optional[str] = None
    received_by_me: float = 0
    received_by_fac: float = 0
    payment_by_me: float = 0
    payment_by_fac: float = 0
    party: str
    mode: str
    note: Optional[str] = ""


class Payment(PaymentBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


# ─── Cash Book (P0 canonical layer) ─────────────────────────────────────────
# Cash Book is a UNIFIED TIMELINE view of every money movement, sourced from
# canonical modules. It NEVER creates customer or vendor payments — those must
# originate from their natural module (Sales Payments / Purchase Payments).
# The `cash_book_entries` collection stores only the transactions that legitimately
# originate from the Cash Book itself: general income, general expense, and
# inter-account transfers.
#
# `source` values:
#   - "cash_book"       — created by user via new Cash Book UI (canonical)
#   - "legacy_shim"     — created via deprecated POST /api/payments; excluded
#                          from all financial KPIs.
#   - "legacy_migrated" — auto-stamped on pre-existing db.payments rows for the
#                          Migration surface in Cash Book timeline (read-only,
#                          excluded from KPIs).

CashBookKind = Literal["general_income", "general_expense", "transfer"]


class CashBookEntryBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    date: Optional[str] = None
    kind: CashBookKind = "general_expense"
    amount: float = 0
    mode: str = "Cash"

    # For general_income / general_expense
    account_id: Optional[str] = ""
    account_name: Optional[str] = ""
    party_name: Optional[str] = ""   # optional counterparty label (e.g. "Airtel")

    # For transfers
    from_account_id: Optional[str] = ""
    from_account_name: Optional[str] = ""
    to_account_id: Optional[str] = ""
    to_account_name: Optional[str] = ""

    reference: Optional[str] = ""
    notes: Optional[str] = ""


class CashBookEntry(CashBookEntryBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    source: str = "cash_book"    # "cash_book" | "legacy_shim" | "legacy_migrated"
    reversed: bool = False
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class Customer(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = ""
    address: Optional[str] = ""
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


# ================================================================
# COMPUTATIONS
# ================================================================
def _item_cost(it: dict) -> float:
    return float(sum(it.get(k, 0) or 0 for k in [
        "factory_complete", "factory_glass", "factory_fitting",
        "outside_complete", "outside_glass", "outside_fitting",
    ]))


def _item_factory(it: dict) -> float:
    return float(sum(it.get(k, 0) or 0 for k in [
        "factory_complete", "factory_glass", "factory_fitting",
    ]))


def _item_outside(it: dict) -> float:
    return float(sum(it.get(k, 0) or 0 for k in [
        "outside_complete", "outside_glass", "outside_fitting",
    ]))


def compute_order_aggregates(order: dict) -> dict:
    """Recalculate stored aggregates. Revenue & profit are recognized ONLY on shipped qty.

    Phase 6 · Slice 3: this function is now a THIN ADAPTER over the pure
    domain helpers `order_realized_amounts`, `order_estimated_amounts`,
    `order_unrealized`. Every money value is computed once, in paise, by
    the domain layer; this adapter converts back to display floats and
    stamps every denormalised field on the `order` dict — preserving the
    exact response contract every caller / test / snapshot relies on.

    Only these responsibilities remain in server.py:
      * Item mutation: stamping `qty_shipped` back onto each item (order
        lifecycle bookkeeping — not pure calculation).
      * `status` auto-update based on shipment progress.
      * `last_shipped_date` extraction.
      * Field aliasing (realized_*, revenue_recognized).

    Freight, packing, and event-recorded adjustments (`other_revenue`,
    `other_expense`) are included AS-IS — never proportioned by shipment
    ratio — matching the pre-refactor rule.
    """
    items = order.get("items") or []
    shipments = order.get("shipments") or []

    # ── Domain layer — pure paise-integer calculations ───────────────────
    real = order_realized_amounts(order)
    est = order_estimated_amounts(order)
    unreal = order_unrealized(order)

    # ── Item-level bookkeeping (stamp qty_shipped on each item) ──────────
    shipped_qty_by_item = _order_shipped_qty_by_item(order)
    ordered_qty_total = 0.0
    shipped_qty_total = 0.0
    for it in items:
        iid = it.get("id")
        ordered_qty_total += float(it.get("qty") or 0)
        sq = shipped_qty_by_item.get(iid, 0.0)
        shipped_qty_total += sq
        it["qty_shipped"] = sq

    # Ordered vs shipped product sales — informational totals derived from
    # domain estimates; ordered = full projection, shipped = realized slice.
    ordered_product_sales = from_paise(est["estimated_product_sales_paise"])
    shipped_product_sales = from_paise(real["shipped_product_sales_paise"])

    # ── Convert every paise integer back to display float ────────────────
    factory_cost_total = from_paise(real["factory_cost_realized_paise"])
    outside_cost_total = from_paise(real["outside_cost_realized_paise"])
    other_revenue_total = from_paise(real["other_revenue_total_paise"])
    other_expense_total = from_paise(real["other_expense_total_paise"])
    ship_freight_charged = from_paise(real["ship_freight_charged_paise"])
    ship_freight_paid = from_paise(real["ship_freight_paid_paise"])
    ship_boxes = sum(float((s or {}).get("boxes_shipped") or 0) for s in shipments)

    operating_revenue = from_paise(real["operating_revenue_paise"])
    total_cost = from_paise(real["total_cost_paise"])
    tax_amount = from_paise(real["tax_amount_paise"])
    invoice_total = from_paise(real["invoice_total_paise"])
    net_profit = from_paise(real["net_profit_paise"])

    margin = (net_profit / operating_revenue * 100.0) if operating_revenue else 0
    progress = (shipped_qty_total / ordered_qty_total * 100.0) if ordered_qty_total else 0

    # ── Status auto-update (unchanged from pre-refactor behaviour) ───────
    manual_states = {"Draft", "Confirmed", "Cancelled", "Delivered"}
    cur = order.get("status") or "Confirmed"
    if cur not in manual_states or cur == "Confirmed":
        if shipped_qty_total <= 0:
            order["status"] = "Packed" if cur in ("Packed", "Partially Shipped", "Fully Shipped") else cur
        elif shipped_qty_total + 0.0001 >= ordered_qty_total and ordered_qty_total > 0:
            order["status"] = "Fully Shipped"
        else:
            order["status"] = "Partially Shipped"

    # Latest shipped date for monthly bucketing
    last = max([(s.get("date") or "") for s in shipments], default="") if shipments else ""
    order["last_shipped_date"] = last or order.get("shipped_date")

    # Bug fix (2026-07-22) — derive order-level `shipped_date` from
    # shipments (source of truth). None while partially shipped, set to
    # the date of the shipment that caused cumulative qty to reach ordered
    # qty when fully shipped. Deterministic + idempotent (repeated calls
    # produce no drift). See domain.derive_completion_shipped_date for
    # the exact rule. Auto-backfills historical orders whose stored
    # `shipped_date` was blank despite full shipments.
    order["shipped_date"] = derive_completion_shipped_date(order)

    # ── Stamp every denormalised field (preserving contract) ─────────────
    order["ordered_qty_total"] = ordered_qty_total
    order["shipped_qty_total"] = shipped_qty_total
    order["ordered_product_sales"] = ordered_product_sales
    order["shipped_product_sales"] = shipped_product_sales
    order["factory_cost_total"] = factory_cost_total
    order["outside_cost_total"] = outside_cost_total
    order["other_revenue_total"] = other_revenue_total
    order["other_expense_total"] = other_expense_total
    order["ship_freight_charged_total"] = ship_freight_charged
    order["ship_freight_paid_total"] = ship_freight_paid
    order["ship_boxes_shipped_total"] = ship_boxes
    # keep legacy freight_* on order in sync too for back-compat
    order["freight_charged"] = ship_freight_charged
    order["freight_paid"] = ship_freight_paid
    order["boxes_shipped"] = ship_boxes
    order["operating_revenue"] = operating_revenue
    order["total_cost"] = total_cost
    order["tax_amount"] = tax_amount
    order["invoice_total"] = invoice_total
    order["net_profit"] = net_profit
    order["margin_percent"] = margin
    order["shipment_progress_percent"] = progress
    # legacy compatibility fields expected by dashboard/breakdown
    order["product_sales_total"] = shipped_product_sales

    # ─── Phase 4 — estimated (full-order) revenue + profit and realized aliases.
    # `estimated_*` = what this order will total when every remaining unit is
    # eventually shipped, at the committed rate + already-recorded freight /
    # packing / other adjustments. `realized_*` = same as the shipped-qty
    # values already on the order — kept as explicit fields so the frontend
    # never has to guess which aggregate is "recognized".
    estimated_operating_revenue = from_paise(est["estimated_operating_revenue_paise"])
    estimated_total_cost = from_paise(est["estimated_total_cost_paise"])
    estimated_net_profit = from_paise(est["estimated_net_profit_paise"])
    estimated_margin = (estimated_net_profit / estimated_operating_revenue * 100.0
                        if estimated_operating_revenue else 0)
    order["estimated_factory_cost_total"] = from_paise(est["estimated_factory_cost_paise"])
    order["estimated_outside_cost_total"] = from_paise(est["estimated_outside_cost_paise"])
    order["estimated_operating_revenue"] = estimated_operating_revenue
    order["estimated_total_cost"] = estimated_total_cost
    order["estimated_net_profit"] = estimated_net_profit
    order["estimated_margin_percent"] = estimated_margin
    # Aliases — same numbers, PRD-mandated names for the UI.
    order["realized_revenue"] = operating_revenue
    order["realized_net_profit"] = net_profit
    order["revenue_recognized"] = operating_revenue
    order["unrealized_revenue"] = max(0.0, from_paise(unreal["unrealized_revenue_paise"]))
    order["unrealized_net_profit"] = from_paise(unreal["unrealized_net_profit_paise"])
    return order


async def _recompute_payment_aggregates_for_orders(order_ids: List[str]):
    """Recompute total_received and outstanding_balance for a set of orders,
    using the CustomerPayment.allocations table as source of truth.

    Phase 6 · Slice 4: allocation sums and outstanding derivation are now
    routed through the shared domain helpers `sum_allocations_to_order`
    and `order_outstanding_from_alloc`. All money math is paise-safe.
    Idempotent — safe to re-run any number of times.
    """
    if not order_ids:
        return
    # Fetch every customer payment that references any of these orders — a
    # single round trip, then let the pure domain helper sum allocations
    # per order. Excludes voided/reversed payments via is_customer_payment_active.
    pays = await db.customer_payments.find(
        {"allocations.order_id": {"$in": order_ids}}, {"_id": 0}
    ).to_list(50000)
    for oid in order_ids:
        doc = await db.orders.find_one({"id": oid}, {"_id": 0})
        if not doc:
            continue
        alloc_p = sum_allocations_to_order(pays, oid)
        invoice_p = to_paise(doc.get("invoice_total"))
        # Pre-refactor rule: `outstanding = invoice - total_received` — could
        # go negative on over-payment. We preserve THAT exact behaviour
        # here (the domain's `order_outstanding_from_alloc` clamps to zero,
        # which is only appropriate for the allocation-UI endpoint, not
        # for the stored aggregate).
        outstanding_p = invoice_p - alloc_p
        total_recv = from_paise(alloc_p)
        outstanding = from_paise(outstanding_p)
        # Payment status: only auto if there ARE payments; else preserve user-set.
        # 50-paise (₹0.50) hysteresis matches the pre-refactor rule
        # ("close-enough-to-invoice → Paid"); this is an ORDER-lifecycle
        # rule kept alongside the write, not a domain calculation.
        pstatus = doc.get("payment_status") or "Unpaid"
        if alloc_p > 0:
            pstatus = "Paid" if alloc_p + 50 >= invoice_p else "Partial"
        elif not doc.get("_had_legacy_payment_status"):
            pstatus = "Unpaid"
        await db.orders.update_one({"id": oid}, {"$set": {
            "total_received": total_recv,
            "outstanding_balance": outstanding,
            "payment_status": pstatus,
        }})


COST_CATEGORIES = ("complete", "glass", "fitting")


def _prep_item(raw: dict) -> dict:
    it = OrderItem(**raw).model_dump()
    if not it.get("product_sales"):
        it["product_sales"] = float(it.get("qty") or 0) * float(it.get("rate") or 0)

    # If purchase_sources are provided, they are the single source of truth
    # for this item's cost — recompute legacy factory_*/outside_* sums from them
    # so compute_order_aggregates + dashboard KPIs stay accurate without
    # touching those code paths.
    sources = it.get("purchase_sources") or []
    if sources:
        fc = fg = ff = oc = og = of = 0.0
        for s in sources:
            comp = float(s.get("complete") or 0)
            gla = float(s.get("glass") or 0)
            fit = float(s.get("fitting") or 0)
            if s.get("supplier_id") == FACTORY_SUPPLIER_ID:
                fc += comp; fg += gla; ff += fit
            else:
                oc += comp; og += gla; of += fit
        it["factory_complete"] = round(fc, 2)
        it["factory_glass"]    = round(fg, 2)
        it["factory_fitting"]  = round(ff, 2)
        it["outside_complete"] = round(oc, 2)
        it["outside_glass"]    = round(og, 2)
        it["outside_fitting"]  = round(of, 2)
    return it


def _synthesise_purchase_sources(it: dict) -> list:
    """For orders migrated from the legacy schema (no purchase_sources) — expose
    the existing factory_*/outside_* fields as two implicit rows so the new UI
    can render + edit them. This is READ-only; on save the frontend echoes
    the (possibly-edited) rows back and the model persists them properly."""
    if it.get("purchase_sources"):
        return it["purchase_sources"]
    rows = []
    fc, fg, ff = (float(it.get(k) or 0) for k in ("factory_complete", "factory_glass", "factory_fitting"))
    if fc + fg + ff > 0:
        rows.append({
            "id": f"legacy-factory-{it.get('id')}",
            "supplier_id": FACTORY_SUPPLIER_ID,
            "supplier_name": "Factory",
            "complete": fc, "glass": fg, "fitting": ff,
        })
    oc, og, of = (float(it.get(k) or 0) for k in ("outside_complete", "outside_glass", "outside_fitting"))
    if oc + og + of > 0:
        rows.append({
            "id": f"legacy-outside-{it.get('id')}",
            "supplier_id": "",   # user must pick a vendor before saving
            "supplier_name": "",
            "complete": oc, "glass": og, "fitting": of,
        })
    return rows


async def _resolve_supplier(supplier_id: str, supplier_name: str) -> tuple[str, str]:
    """Return (canonical_supplier_id, canonical_vendor_name) for a purchase row.
    Factory → canonical Father's Firm identity so vendor payables and party
    ledger balance updates route through the FF party."""
    if supplier_id == FACTORY_SUPPLIER_ID:
        return FACTORY_SUPPLIER_ID, FACTORY_PARTY_NAME
    return supplier_id or "", (supplier_name or "").strip()


def _linked_source_key(order_id: str, item_id: str, source_row_id: str, category: str) -> str:
    return f"{order_id}::{item_id}::{source_row_id}::{category}"


async def _sync_order_linked_purchases(order: dict) -> dict:
    """Mirror every non-zero (item, purchase_source, category) tuple into the
    Purchases collection as an auto-managed `order_product_purchase` row.

    - Stable key ⇒ repeated saves NEVER create duplicates.
    - Amount → 0 or row/source removed ⇒ delete the linked purchase, UNLESS it
      already has payments. In that case the purchase is preserved but marked
      `stale=True` so the UI can prompt for a manual adjustment.
    - Factory rows post under the Father's Firm supplier so the FF settlement
      balance updates through the same derived-ledger path as any other vendor.
    """
    order_id = order.get("id")
    if not order_id:
        return {"created": 0, "updated": 0, "deleted": 0, "kept_paid": 0, "errors": []}

    keep_keys: set[str] = set()
    created = updated = deleted = kept_paid = 0
    errors: list[str] = []
    order_date = (order.get("order_date") or order.get("shipped_date") or
                  datetime.now(timezone.utc).date().isoformat())

    for it in (order.get("items") or []):
        for src in (it.get("purchase_sources") or []):
            supplier_id_raw = (src.get("supplier_id") or "").strip()
            supplier_name_raw = (src.get("supplier_name") or "").strip()
            if not supplier_id_raw and not supplier_name_raw:
                # Blank supplier — validation error only if there's a non-zero amount.
                if any(float(src.get(c) or 0) > 0.005 for c in COST_CATEGORIES):
                    errors.append(f"Item '{it.get('product_name') or it.get('id')}' has a purchase row without a supplier.")
                continue
            supplier_id, vendor_name = await _resolve_supplier(supplier_id_raw, supplier_name_raw)
            if not vendor_name:
                continue
            # Bug fix (2026-07-22): resolve the canonical vendor party ID
            # once per (supplier_id, vendor_name) tuple so both new and
            # updated linked purchases carry it. Factory rows route to
            # SYSTEM_FF_ID via the FF alias check; other rows resolve
            # via the deterministic get_or_create_vendor_party helper.
            vendor_party_id: str | None = None
            if supplier_id == FACTORY_SUPPLIER_ID or is_ff_alias(vendor_name):
                vendor_party_id = SYSTEM_FF_ID
            elif vendor_name:
                # get_or_create_vendor_party is idempotent — it returns the
                # existing canonical party if one already matches, else
                # creates a new one. Never guesses on ambiguous matches.
                party = await get_or_create_vendor_party(db, vendor_name)
                if party and party.get("id"):
                    vendor_party_id = party["id"]

            for cat in COST_CATEGORIES:
                amt = float(src.get(cat) or 0)
                if amt <= 0.005:
                    continue
                key = _linked_source_key(order_id, it.get("id") or "", src.get("id") or "", cat)
                keep_keys.add(key)

                pur_item = {
                    "id": str(uuid.uuid4()),
                    "category": cat.title(),  # Complete / Glass / Fitting
                    "description": f"{it.get('product_name') or 'Product'} — {cat.title()}",
                    "qty": 1, "rate": amt, "amount": amt,
                }
                base_doc = {
                    "vendor_name": vendor_name,
                    "purchase_date": order_date,
                    "items": [pur_item],
                    "freight": 0, "other_charges": 0,
                    "tax_applicable": False, "tax_type": "None", "tax_percent": 0,
                    "tax_amount": 0, "tax_amount_manual": False,
                    "notes": f"Auto-linked from Order {order_id[:8]} · Item {(it.get('product_name') or '').strip()}",
                    "linked_to_order_id": order_id,
                    "linked_source_key": key,
                    "linked_supplier_id": supplier_id,
                    "linked_source_row_id": src.get("id"),
                    "linked_order_item_id": it.get("id"),
                    "linked_cost_category": cat,
                    "source_type": "order_product_purchase",
                    # Bug fix (2026-07-22): stamp the resolved canonical
                    # vendor party ID so vendor payables, Party Ledger v2,
                    # and reconciliation all key off the stable ID rather
                    # than the (renamable, duplicable) display name.
                    "vendor_party_id": vendor_party_id,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }

                existing = await db.purchases.find_one({"linked_source_key": key}, {"_id": 0})
                if existing:
                    total_paid = float(existing.get("total_paid") or 0)
                    invoice_total = amt
                    outstanding = round(invoice_total - total_paid, 2)
                    payment_status = "Paid" if total_paid + 0.005 >= invoice_total \
                                    else ("Partial" if total_paid > 0.005 else "Unpaid")
                    upd = {
                        **base_doc,
                        "subtotal": amt, "invoice_total": amt,
                        "outstanding_balance": outstanding,
                        "payment_status": payment_status,
                        "stale": False,
                    }
                    # Preserve identity + payment history
                    upd["items"][0]["id"] = (existing.get("items") or [{}])[0].get("id") or upd["items"][0]["id"]
                    upd["id"] = existing.get("id")
                    await db.purchases.update_one({"id": existing["id"]}, {"$set": upd})
                    updated += 1
                else:
                    doc = {
                        **base_doc,
                        "subtotal": amt, "invoice_total": amt,
                        "total_paid": 0, "outstanding_balance": amt,
                        "payment_status": "Unpaid",
                        "created_at": datetime.now(timezone.utc).isoformat(),
                        "stale": False,
                    }
                    doc["id"] = str(uuid.uuid4())
                    await db.purchases.insert_one(doc)
                    # Ensure vendor exists in vendor directory (skip for Factory)
                    if supplier_id != FACTORY_SUPPLIER_ID and vendor_name:
                        await db.vendors.update_one(
                            {"name": vendor_name},
                            {"$setOnInsert": Vendor(name=vendor_name).model_dump()},
                            upsert=True,
                        )
                    created += 1

    # Purge / mark stale any linked purchases for this order that are no longer needed
    async for old in db.purchases.find(
        {"linked_to_order_id": order_id, "source_type": "order_product_purchase"}, {"_id": 0}
    ):
        if old.get("linked_source_key") in keep_keys:
            continue
        total_paid = float(old.get("total_paid") or 0)
        if total_paid > 0.005:
            await db.purchases.update_one({"id": old["id"]}, {"$set": {
                "stale": True,
                "notes": (old.get("notes") or "") + " · REMOVED FROM ORDER — has payments, needs adjustment.",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }})
            kept_paid += 1
        else:
            await db.purchases.delete_one({"id": old["id"]})
            deleted += 1

    return {"created": created, "updated": updated, "deleted": deleted,
            "kept_paid": kept_paid, "errors": errors}


async def _delete_order_linked_purchases(order_id: str) -> dict:
    """Called when an order is deleted — same policy as sync but for full removal:
    delete unpaid linked purchases, keep the ones with payments as stale."""
    deleted = kept_paid = 0
    async for old in db.purchases.find(
        {"linked_to_order_id": order_id, "source_type": "order_product_purchase"}, {"_id": 0}
    ):
        total_paid = float(old.get("total_paid") or 0)
        if total_paid > 0.005:
            await db.purchases.update_one({"id": old["id"]}, {"$set": {
                "stale": True,
                "notes": (old.get("notes") or "") + " · SOURCE ORDER DELETED — paid; kept for adjustment.",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }})
            kept_paid += 1
        else:
            await db.purchases.delete_one({"id": old["id"]})
            deleted += 1
    return {"deleted": deleted, "kept_paid": kept_paid}


# ================================================================
# ORDERS
# ================================================================
def _validate_purchase_sources(order: dict) -> list[str]:
    """Fail-fast validation used BEFORE the order is inserted / updated so a
    validation error can never leave a half-written document behind."""
    errors: list[str] = []
    for it in (order.get("items") or []):
        for src in (it.get("purchase_sources") or []):
            has_amount = any(float(src.get(c) or 0) > 0.005 for c in COST_CATEGORIES)
            has_supplier = bool((src.get("supplier_id") or "").strip() or (src.get("supplier_name") or "").strip())
            if has_amount and not has_supplier:
                errors.append(
                    f"Item '{it.get('product_name') or it.get('id')}' has a purchase row without a supplier."
                )
    return errors


@api_router.post("/orders", response_model=Order)
async def create_order(payload: OrderBase):
    data = payload.model_dump()
    data["items"] = [_prep_item(i) for i in data.get("items", [])]
    validation_errors = _validate_purchase_sources(data)
    if validation_errors:
        raise HTTPException(400, {"detail": "Some purchase rows are missing a supplier.",
                                  "errors": validation_errors})
    order = Order(**data).model_dump()
    compute_order_aggregates(order)
    order["updated_at"] = datetime.now(timezone.utc).isoformat()
    # Phase 2: stamp stable customer party id onto the order.
    party = await get_or_create_customer_party(db, data.get("client_name") or "")
    if party:
        order["customer_party_id"] = party["id"]
    await db.orders.insert_one(order)
    # Ensure customer exists
    if data.get("client_name"):
        await db.customers.update_one(
            {"name": data["client_name"]},
            {"$setOnInsert": Customer(name=data["client_name"]).model_dump()},
            upsert=True,
        )
    await _sync_order_linked_purchases(order)
    return Order(**order)


@api_router.get("/orders", response_model=List[Order])
async def list_orders(
    client_name: Optional[str] = None,
    payment_status: Optional[str] = None,
    main_category: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    q: dict = {}
    if client_name:
        q["client_name"] = {"$regex": client_name, "$options": "i"}
    if payment_status and payment_status != "all":
        q["payment_status"] = payment_status
    if main_category and main_category != "all":
        q["items.main_category"] = main_category
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        q["shipped_date"] = d
    docs = await db.orders.find(q, {"_id": 0}).sort([("last_shipped_date", -1), ("shipped_date", -1)]).to_list(5000)
    for d in docs:
        for it in (d.get("items") or []):
            if not it.get("purchase_sources"):
                it["purchase_sources"] = _synthesise_purchase_sources(it)
    return [Order(**d) for d in docs]


@api_router.get("/orders/{oid}", response_model=Order)
async def get_order(oid: str):
    d = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Order not found")
    for it in (d.get("items") or []):
        if not it.get("purchase_sources"):
            it["purchase_sources"] = _synthesise_purchase_sources(it)
    return Order(**d)


@api_router.put("/orders/{oid}", response_model=Order)
async def update_order(oid: str, payload: OrderBase):
    existing = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Order not found")
    data = payload.model_dump()
    data["items"] = [_prep_item(i) for i in data.get("items", [])]
    validation_errors = _validate_purchase_sources(data)
    if validation_errors:
        raise HTTPException(400, {"detail": "Some purchase rows are missing a supplier.",
                                  "errors": validation_errors})
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    compute_order_aggregates(data)
    # Phase 2: preserve customer_party_id when the client name matches or
    # resolves to the same party; changing to a different customer only
    # updates the id when the caller has explicitly reassigned via a fresh
    # client_name (party resolver picks the correct existing/new id).
    party = await get_or_create_customer_party(db, data.get("client_name") or "")
    if party:
        data["customer_party_id"] = party["id"]
    await db.orders.update_one({"id": oid}, {"$set": data})
    updated = await db.orders.find_one({"id": oid}, {"_id": 0})
    await _sync_order_linked_purchases(updated)
    return Order(**updated)


@api_router.delete("/orders/{oid}")
async def delete_order(oid: str):
    existing = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Order not found")
    linked_report = await _delete_order_linked_purchases(oid)
    if linked_report["kept_paid"] > 0:
        raise HTTPException(400, {
            "detail": f"{linked_report['kept_paid']} linked purchase(s) already have payments. "
                      "Reverse those payments before deleting the order.",
            "kept_paid": linked_report["kept_paid"],
        })
    await db.orders.delete_one({"id": oid})
    return {"deleted": True, "linked_purchases_removed": linked_report["deleted"]}


# ================================================================
# PAYMENTS (unchanged from before)
# ================================================================
# ================================================================
# LEGACY /payments endpoints — DEPRECATED SHIM
# ================================================================
# These endpoints once wrote directly to `db.payments` (the pre-refactor Cash
# Book). Cash Book is now a READ-ONLY unified timeline; genuine income /
# expense / transfer must flow through POST /api/cash-book-entries.
#
# For strict back-compat with any external caller still hitting these routes,
# we accept the write, coerce the shape into a CashBookEntry stamped
# source='legacy_shim', and return the classic Payment payload. Shim rows are
# EXCLUDED from every canonical KPI so a stale integration cannot double-count.

def _legacy_payment_to_cbe(p: dict) -> dict:
    received = float((p.get("received_by_me") or 0) + (p.get("received_by_fac") or 0))
    paid = float((p.get("payment_by_me") or 0) + (p.get("payment_by_fac") or 0))
    if received >= paid:
        kind = "general_income"
        amount = received - paid
    else:
        kind = "general_expense"
        amount = paid - received
    return CashBookEntry(
        date=p.get("date"),
        kind=kind if amount > 0 else "general_expense",
        amount=amount,
        mode=p.get("mode") or "Cash",
        party_name=p.get("party") or "",
        notes=p.get("note") or "",
        source="legacy_shim",
    ).model_dump()


@api_router.post("/payments", response_model=Payment, deprecated=True)
async def create_payment(payload: PaymentBase):
    """DEPRECATED: use POST /api/cash-book-entries. Kept for API back-compat only.

    Rows created here are stamped `source='legacy_shim'` and NEVER counted in
    dashboard KPIs / exports."""
    p = Payment(**payload.model_dump()).model_dump()
    # Store in legacy collection (unchanged) so GET /payments continues to work.
    await db.payments.insert_one({**p, "source": "legacy_shim"})
    # Also mirror to cash_book_entries as a legacy_shim entry for the unified
    # timeline (still excluded from KPIs).
    cbe = _legacy_payment_to_cbe(p)
    cbe["notes"] = (cbe.get("notes") or "") + f" [shim of legacy payment {p['id'][:8]}]"
    cbe["reference"] = f"legacy:{p['id']}"
    await db.cash_book_entries.insert_one(cbe)
    return Payment(**p)


@api_router.get("/payments", response_model=List[Payment])
async def list_payments(
    party: Optional[str] = None,
    mode: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """Read-only legacy view. Returns pre-existing db.payments rows (stamped
    `source='legacy_migrated'`) plus any shim rows. Excluded from KPIs."""
    q: dict = {}
    if party:
        q["party"] = {"$regex": party, "$options": "i"}
    if mode and mode != "all":
        q["mode"] = mode
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        q["date"] = d
    docs = await db.payments.find(q, {"_id": 0}).sort("date", 1).to_list(5000)
    return [Payment(**{k: v for k, v in d.items() if k in Payment.model_fields})
            for d in docs]


@api_router.put("/payments/{pid}", response_model=Payment, deprecated=True)
async def update_payment(pid: str, payload: PaymentBase):
    """DEPRECATED: legacy Cash Book rows can no longer be created; they can
    only be edited in place. Prefer editing the origin document."""
    existing = await db.payments.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Payment not found")
    await db.payments.update_one({"id": pid}, {"$set": payload.model_dump()})
    updated = await db.payments.find_one({"id": pid}, {"_id": 0})
    return Payment(**{k: v for k, v in updated.items() if k in Payment.model_fields})


@api_router.delete("/payments/{pid}", deprecated=True)
async def delete_payment(pid: str):
    """DEPRECATED. Legacy row removal — no side effects on canonical modules."""
    res = await db.payments.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Payment not found")
    # Also purge any shim mirror
    await db.cash_book_entries.delete_many({"reference": f"legacy:{pid}"})
    return {"deleted": True}


# ================================================================
# CUSTOMERS
# ================================================================
@api_router.get("/orders/{oid}/payments")
async def order_payments_for(oid: str):
    """List every customer payment that has an allocation to THIS order.
    Also returns the parent payment's account/mode/reference and this customer's
    unallocated advance so the OrderDialog can display receipts + one-click
    advance allocation without duplicating payment records."""
    order = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    pays = await db.customer_payments.find(
        {"allocations.order_id": oid}, {"_id": 0}
    ).sort("date", 1).to_list(2000)
    rows = []
    for p in pays:
        alloc = next((a for a in (p.get("allocations") or []) if a.get("order_id") == oid), None)
        if not alloc:
            continue
        pay_amt = float(p.get("amount") or 0)
        alloc_amt = float(alloc.get("amount") or 0)
        # Rough per-order status derived from the parent payment
        if alloc_amt + 0.5 >= pay_amt:
            pstatus = "Full"
        elif alloc_amt > 0.5:
            pstatus = "Partial"
        else:
            pstatus = "Advance"
        rows.append({
            "payment_id": p.get("id"),
            "date": p.get("date"),
            "customer_name": p.get("customer_name"),
            "allocated_to_this_order": alloc_amt,
            "total_amount": pay_amt,
            "mode": p.get("mode"),
            "account_id": p.get("account_id"),
            "account_name": p.get("account_name"),
            "reference": p.get("reference"),
            "remarks": p.get("remarks"),
            "received_by_party_name": p.get("received_by_party_name"),
            "payment_status": pstatus,
        })
    total_received = sum(r["allocated_to_this_order"] for r in rows)
    invoice_total = float(order.get("invoice_total") or 0)

    # Any unallocated advance from this same customer — pointer to the oldest
    # advance payment so the UI can call POST /allocate-advance to top-up.
    cust = order.get("client_name") or ""
    advance_total = 0.0
    advance_payment_id = None
    if cust:
        adv_docs = await db.customer_payments.find(
            {"customer_name": cust, "unallocated": {"$gt": 0.5}}, {"_id": 0}
        ).sort("date", 1).to_list(200)
        for a in adv_docs:
            advance_total += float(a.get("unallocated") or 0)
        if adv_docs:
            advance_payment_id = adv_docs[0].get("id")

    return {
        "order_id": oid,
        "customer_name": cust,
        "invoice_total": invoice_total,
        "total_received": round(total_received, 2),
        "outstanding": round(invoice_total - total_received, 2),
        "customer_advance_available": round(advance_total, 2),
        "advance_payment_id": advance_payment_id,
        "count": len(rows),
        "payments": rows,
    }


class AllocateAdvanceIn(BaseModel):
    amount: Optional[float] = None


@api_router.post("/orders/{oid}/allocate-advance")
async def allocate_customer_advance_to_order(oid: str, body: Optional[AllocateAdvanceIn] = None):
    """Reduce this customer's unallocated advance and top-up the allocation to THIS
    order — WITHOUT creating a new payment record. If `amount` is omitted, it uses
    min(advance_available, outstanding_on_order). Consumes advances FIFO."""
    amount = body.amount if body is not None else None
    order = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    cust = order.get("client_name") or ""
    if not cust:
        raise HTTPException(400, "Order has no customer")

    invoice = float(order.get("invoice_total") or 0)
    total_recv = float(order.get("total_received") or 0)
    outstanding = max(0.0, invoice - total_recv)

    adv_docs = await db.customer_payments.find(
        {"customer_name": cust, "unallocated": {"$gt": 0.5}}, {"_id": 0}
    ).sort("date", 1).to_list(200)
    if not adv_docs:
        raise HTTPException(400, "No advance available for this customer")

    total_advance = sum(float(p.get("unallocated") or 0) for p in adv_docs)
    if outstanding <= 0.5:
        raise HTTPException(400, "This order has no outstanding balance")

    want = float(amount) if amount is not None else min(total_advance, outstanding)
    want = max(0.0, min(want, total_advance, outstanding))
    if want <= 0.5:
        raise HTTPException(400, "Nothing to allocate")

    allocated_from_advances = 0.0
    touched_payment_ids: List[str] = []
    for p in adv_docs:
        if allocated_from_advances + 0.001 >= want:
            break
        avail = float(p.get("unallocated") or 0)
        take = min(avail, want - allocated_from_advances)
        if take <= 0.001:
            continue
        # Update existing allocation on this payment or push a new one
        allocs = list(p.get("allocations") or [])
        found = False
        for a in allocs:
            if a.get("order_id") == oid:
                a["amount"] = float(a.get("amount") or 0) + take
                found = True
                break
        if not found:
            allocs.append({"order_id": oid, "amount": take})
        pay_amt = float(p.get("amount") or 0)
        alloc_total = sum(float((a or {}).get("amount") or 0) for a in allocs)
        await db.customer_payments.update_one(
            {"id": p["id"]},
            {"$set": {
                "allocations": allocs,
                "allocated_total": round(alloc_total, 2),
                "unallocated": round(max(0.0, pay_amt - alloc_total), 2),
            }},
        )
        touched_payment_ids.append(p["id"])
        allocated_from_advances += take

    await _recompute_payment_aggregates_for_orders([oid])
    return {
        "allocated": round(allocated_from_advances, 2),
        "order_id": oid,
        "touched_payment_ids": touched_payment_ids,
    }


@api_router.get("/orders/{oid}/timeline")
async def order_timeline(oid: str):
    """Chronological list of events for this order — order creation, shipments,
    customer payments (allocations), and status changes. Frontend-friendly."""
    order = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")

    events: List[dict] = []

    created = order.get("order_date") or (order.get("created_at") or "")[:10]
    events.append({
        "type": "order_created",
        "date": created,
        "title": "Order Confirmed",
        "detail": f"{len(order.get('items') or [])} items · {float(order.get('ordered_qty_total') or 0):.0f} pcs",
    })

    for i, sh in enumerate(sorted((order.get("shipments") or []), key=lambda s: s.get("date") or "")):
        qty = sum(float((si or {}).get("qty") or 0) for si in (sh.get("items") or []))
        events.append({
            "type": "shipment",
            "date": sh.get("date"),
            "title": f"Shipment {i + 1}",
            "detail": f"{qty:.0f} pcs · {sh.get('transporter') or 'Transport pending'}"
                      + (f" · LR {sh.get('lr_number')}" if sh.get("lr_number") else ""),
            "shipment_id": sh.get("id"),
        })

    pays = await db.customer_payments.find(
        {"allocations.order_id": oid}, {"_id": 0}
    ).sort("date", 1).to_list(2000)
    for p in pays:
        alloc = next((a for a in (p.get("allocations") or []) if a.get("order_id") == oid), None)
        if not alloc:
            continue
        events.append({
            "type": "payment",
            "date": p.get("date"),
            "title": f"Customer Payment · {p.get('mode') or 'Cash'}",
            "detail": f"₹{float(alloc.get('amount') or 0):,.0f} allocated"
                      + (f" · via {p.get('received_by_party_name')}" if p.get("received_by_party_name") else "")
                      + (f" · Ref {p.get('reference')}" if p.get("reference") else ""),
            "payment_id": p.get("id"),
        })

    events.sort(key=lambda e: (e.get("date") or ""))
    return {"order_id": oid, "count": len(events), "events": events}


@api_router.get("/customers/{name}/outstanding-orders")
async def customer_outstanding_orders(name: str):
    """Return this customer's orders with an outstanding balance, oldest first — used
    to allocate a new payment across invoices.

    Phase 6 · Slice 4: money math paise-safe. Behaviour preserved — the
    display outstanding is CLAMPED to zero (over-payments show as 0 in
    the allocation UI), while the stored `outstanding_balance` may be
    negative for over-paid orders.
    """
    orders = await db.orders.find({"client_name": name}, {"_id": 0}).to_list(5000)
    rows = []
    for o in orders:
        outstanding_p = to_paise(o.get("outstanding_balance"))
        # 50-paise hysteresis (matches pre-refactor `> 0.5`).
        if outstanding_p > 50 or (o.get("payment_status") in ("Unpaid", "Partial")):
            rows.append({
                "id": o.get("id"),
                "date": o.get("last_shipped_date") or o.get("shipped_date") or o.get("order_date"),
                "invoice_total": from_paise(to_paise(o.get("invoice_total"))),
                "total_received": from_paise(to_paise(o.get("total_received"))),
                "outstanding": from_paise(max(0, outstanding_p)),
                "status": o.get("status"),
                "payment_status": o.get("payment_status"),
                "short_id": (o.get("id") or "")[:8],
            })
    rows.sort(key=lambda r: r.get("date") or "")
    return {"count": len(rows), "orders": rows}


@api_router.get("/customers", response_model=List[Customer])
async def list_customers():
    docs = await db.customers.find({}, {"_id": 0}).sort("name", 1).to_list(5000)
    return [Customer(**d) for d in docs]


# ================================================================
# DASHBOARD
# ================================================================
@api_router.get("/dashboard")
async def dashboard():
    orders = await db.orders.find({}, {"_id": 0}).to_list(10000)

    operating_revenue = sum(o.get("operating_revenue") or 0 for o in orders)
    invoice_value = sum(o.get("invoice_total") or 0 for o in orders)
    total_cost = sum(o.get("total_cost") or 0 for o in orders)
    net_profit = sum(o.get("net_profit") or 0 for o in orders)
    gst_collected = sum(o.get("tax_amount") or 0 for o in orders)
    margin = (net_profit / operating_revenue * 100.0) if operating_revenue else 0

    # Phase 4 — estimated (full-order) revenue + profit aggregates.
    # Falls back to realized (operating_revenue / net_profit) for orders that
    # pre-date the Phase 4 backfill and don't yet have the new fields.
    estimated_revenue = sum((o.get("estimated_operating_revenue")
                             if o.get("estimated_operating_revenue") is not None
                             else o.get("operating_revenue") or 0) for o in orders)
    estimated_total_cost = sum((o.get("estimated_total_cost")
                                if o.get("estimated_total_cost") is not None
                                else o.get("total_cost") or 0) for o in orders)
    estimated_net_profit = sum((o.get("estimated_net_profit")
                                if o.get("estimated_net_profit") is not None
                                else o.get("net_profit") or 0) for o in orders)
    estimated_margin = (estimated_net_profit / estimated_revenue * 100.0
                        if estimated_revenue else 0)
    unrealized_revenue = max(0.0, estimated_revenue - operating_revenue)
    unrealized_net_profit = estimated_net_profit - net_profit

    # ─── P0: KPIs `received` / `paid` / `modes` come from CANONICAL sources.
    # Legacy db.payments is intentionally NOT read here — those rows are
    # stamped source='legacy_migrated' and surfaced only in the Cash Book
    # timeline as read-only "Migration" entries.
    # P3: Transfers (both `db.transfers` and any cash_book_entries[kind=transfer]
    # not yet migrated) are ALSO excluded — transfers are profit-neutral.
    #
    # Phase 6 · Slice 2: active-record filtering is now the responsibility
    # of the domain layer. This endpoint fetches EVERYTHING (no inline
    # Mongo filter) and lets `is_*_active` / `is_cash_book_entry_canonical`
    # decide what's canonical. Guarantees the same rule reconcile applies.
    cust_pays = await db.customer_payments.find({}, {"_id": 0}).to_list(20000)
    purchase_pays = await db.purchase_payments.find({}, {"_id": 0}).to_list(20000)
    cb_entries_all = await db.cash_book_entries.find({}, {"_id": 0}).to_list(20000)

    received = from_paise(sum_received_kpi(cust_pays, cb_entries_all))
    paid = from_paise(sum_paid_kpi(purchase_pays, cb_entries_all))

    # Outstanding receivable (Bug fix 2026-07-22): must sum the REMAINING
    # order balance (invoice_total − allocated payments, clamped to zero
    # per order), NOT the full invoice value of unpaid/partial orders.
    # A ₹96,300 order with ₹75,000 paid should contribute ₹21,300, not
    # ₹96,300. Cancelled orders excluded. Single-sourced with
    # /api/dashboard/breakdown.receivable via the shared domain helper.
    outstanding_receivable = from_paise(
        sum_dashboard_outstanding_receivable_paise(orders)
    )
    outstanding_payable = paid  # kept simple — historical shape

    # Boxes and freight
    boxes_used = sum(o.get("boxes_used") or 0 for o in orders)
    boxes_shipped = sum(o.get("boxes_shipped") or 0 for o in orders)
    freight_charged = sum(o.get("freight_charged") or 0 for o in orders)
    freight_paid = sum(o.get("freight_paid") or 0 for o in orders)
    packing_cost = sum(o.get("packing_cost") or 0 for o in orders)

    # Customer advances (unallocated portion of active customer payments)
    customer_advances = from_paise(
        compute_party_metrics(cust_pays, purchase_pays)["customer_advances_paise"]
    )

    # Purchase KPIs (vendor bills & payments). The `purchase_paid` KPI must
    # equal Σ(active purchase_payments.amount) — routed via domain.
    purchases = await db.purchases.find({}, {"_id": 0}).to_list(20000)
    purchase_value = sum((p.get("invoice_total") or 0) for p in purchases)
    purchase_paid = from_paise(sum(
        to_paise(p.get("amount")) for p in purchase_pays
        if is_purchase_payment_active(p)
    ))
    purchase_outstanding = sum(
        (p.get("outstanding_balance") or 0) for p in purchases
        if p.get("payment_status") in ("Unpaid", "Partial")
    )

    # Monthly (by shipped_date)
    monthly = defaultdict(lambda: {"revenue": 0, "profit": 0, "cost": 0})
    for o in orders:
        d = o.get("shipped_date")
        if not d:
            continue
        try:
            dt = datetime.fromisoformat(d.replace("Z", ""))
            key = dt.strftime("%Y-%m")
        except Exception:
            continue
        monthly[key]["revenue"] += o.get("operating_revenue") or 0
        monthly[key]["profit"] += o.get("net_profit") or 0
        monthly[key]["cost"] += o.get("total_cost") or 0
    monthly_series = [{"month": k, **v} for k, v in sorted(monthly.items())]

    # By main category (over items)
    main_cat = defaultdict(lambda: {"sales": 0, "profit_share": 0, "count": 0})
    # profit share requires allocation of order-level costs proportionally
    for o in orders:
        rev = o.get("operating_revenue") or 0
        prof = o.get("net_profit") or 0
        item_sales_sum = sum(i.get("product_sales") or 0 for i in o.get("items") or [])
        for it in o.get("items") or []:
            mc = it.get("main_category") or "Uncategorised"
            main_cat[mc]["sales"] += it.get("product_sales") or 0
            main_cat[mc]["count"] += 1
            share = 0
            if item_sales_sum > 0:
                share = prof * (it.get("product_sales") or 0) / item_sales_sum
            main_cat[mc]["profit_share"] += share
    main_categories = sorted(
        [{"main_category": k, **v} for k, v in main_cat.items()],
        key=lambda x: -x["profit_share"],
    )

    # Sub category drill-down within main
    sub_cat = defaultdict(lambda: defaultdict(lambda: {"sales": 0, "profit_share": 0, "count": 0}))
    for o in orders:
        prof = o.get("net_profit") or 0
        item_sales_sum = sum(i.get("product_sales") or 0 for i in o.get("items") or [])
        for it in o.get("items") or []:
            mc = it.get("main_category") or "Uncategorised"
            sc = it.get("sub_category") or "—"
            sub_cat[mc][sc]["sales"] += it.get("product_sales") or 0
            sub_cat[mc][sc]["count"] += 1
            if item_sales_sum > 0:
                sub_cat[mc][sc]["profit_share"] += prof * (it.get("product_sales") or 0) / item_sales_sum
    sub_categories = {
        mc: sorted([{"sub_category": s, **v} for s, v in subs.items()],
                   key=lambda x: -x["sales"])
        for mc, subs in sub_cat.items()
    }

    # Top customers by profit
    cust = defaultdict(lambda: {"revenue": 0, "profit": 0, "orders": 0})
    for o in orders:
        c = o.get("client_name") or "Unknown"
        cust[c]["revenue"] += o.get("operating_revenue") or 0
        cust[c]["profit"] += o.get("net_profit") or 0
        cust[c]["orders"] += 1
    top_customers = sorted(
        [{"client": k, **v} for k, v in cust.items()],
        key=lambda x: -x["profit"],
    )[:10]

    # Top products by sales (item-level)
    prod = defaultdict(lambda: {"sales": 0, "qty": 0, "orders": 0, "main_category": ""})
    for o in orders:
        for it in o.get("items") or []:
            pkey = f"{it.get('main_category','')}/{it.get('product_name','')}"
            prod[pkey]["sales"] += it.get("product_sales") or 0
            prod[pkey]["qty"] += it.get("qty") or 0
            prod[pkey]["orders"] += 1
            prod[pkey]["main_category"] = it.get("main_category", "")
    top_products = sorted(
        [{"product": k.split("/", 1)[-1], **v} for k, v in prod.items()],
        key=lambda x: -x["sales"],
    )[:10]

    # Payments by mode — sourced from canonical modules only (transfers excluded).
    # Phase 6 · Slice 2: bucketing is now the domain layer's job. The
    # sentinel key "" (blank/None mode) is preserved as an explicit
    # "Other" bucket per the Slice 2 spec — no transaction is silently dropped.
    mode_totals = sum_mode_totals(cust_pays, purchase_pays, cb_entries_all)
    mode_series = [
        {
            "mode": (mode_key if mode_key else "Other"),
            "received": from_paise(vals["received_paise"]),
            "paid": from_paise(vals["paid_paise"]),
        }
        for mode_key, vals in mode_totals.items()
    ]

    return {
        "kpis": {
            "operating_revenue": operating_revenue,
            "invoice_value": invoice_value,
            "total_cost": total_cost,
            "net_profit": net_profit,
            "margin_percent": margin,
            # Phase 4 — estimated + realized split
            "estimated_revenue": estimated_revenue,
            "estimated_total_cost": estimated_total_cost,
            "estimated_net_profit": estimated_net_profit,
            "estimated_margin_percent": estimated_margin,
            "realized_revenue": operating_revenue,          # alias
            "realized_net_profit": net_profit,              # alias
            "revenue_recognized": operating_revenue,        # PRD name
            "unrealized_revenue": unrealized_revenue,
            "unrealized_net_profit": unrealized_net_profit,
            "gst_collected": gst_collected,
            "outstanding_receivable": outstanding_receivable,
            "outstanding_payable": outstanding_payable,
            "received": received,
            "paid": paid,
            "order_count": len(orders),
            "boxes_used": boxes_used,
            "boxes_shipped": boxes_shipped,
            "freight_charged": freight_charged,
            "freight_paid": freight_paid,
            "packing_cost": packing_cost,
            "customer_advances": customer_advances,
            "purchase_value": purchase_value,
            "purchase_paid": purchase_paid,
            "purchase_outstanding": purchase_outstanding,
            "purchase_count": len(purchases),
        },
        "monthly": monthly_series,
        "main_categories": main_categories,
        "sub_categories": sub_categories,
        "top_customers": top_customers,
        "top_products": top_products,
        "modes": mode_series,
    }


# ================================================================
# DASHBOARD BREAKDOWN (drill-down data for interactive KPIs)
# ================================================================
@api_router.get("/dashboard/breakdown")
async def dashboard_breakdown():
    orders = await db.orders.find({}, {"_id": 0}).to_list(10000)

    # ---- Revenue ----
    product_sales_total = sum(o.get("product_sales_total") or 0 for o in orders)
    freight_charged_total = sum(o.get("freight_charged") or 0 for o in orders)
    packing_recovery_total = sum(o.get("packing_recovery") or 0 for o in orders)
    other_revenue_total = sum(o.get("other_revenue_total") or 0 for o in orders)

    # collect other-revenue entries grouped by description
    other_rev_by_desc = defaultdict(lambda: {"amount": 0, "count": 0})
    for o in orders:
        for e in (o.get("other_revenue") or []):
            k = (e.get("description") or "Unlabelled").strip() or "Unlabelled"
            other_rev_by_desc[k]["amount"] += from_paise(to_paise(e.get("amount")))
            other_rev_by_desc[k]["count"] += 1

    rev_by_main = defaultdict(float)
    rev_sub_by_main = defaultdict(lambda: defaultdict(float))
    for o in orders:
        for it in o.get("items") or []:
            mc = it.get("main_category") or "Uncategorised"
            sc = it.get("sub_category") or "—"
            rev_by_main[mc] += it.get("product_sales") or 0
            rev_sub_by_main[mc][sc] += it.get("product_sales") or 0

    revenue = {
        "product_sales": product_sales_total,
        "freight_charged": freight_charged_total,
        "packing_charged": packing_recovery_total,
        "other_revenue": other_revenue_total,
        "total": product_sales_total + freight_charged_total + packing_recovery_total + other_revenue_total,
        "by_main_category": sorted(
            [{"main_category": k, "amount": v} for k, v in rev_by_main.items()],
            key=lambda x: -x["amount"],
        ),
        "by_sub_category": {
            mc: sorted([{"sub_category": s, "amount": a} for s, a in subs.items()],
                       key=lambda x: -x["amount"])
            for mc, subs in rev_sub_by_main.items()
        },
        "other_revenue_by_description": sorted(
            [{"description": k, **v} for k, v in other_rev_by_desc.items()],
            key=lambda x: -x["amount"],
        ),
    }

    # ---- Invoice / Tax ----
    tax_buckets = defaultdict(lambda: {"count": 0, "tax_amount": 0, "invoice_total": 0, "revenue": 0})
    non_tax_revenue = 0
    for o in orders:
        if o.get("tax_applicable"):
            key = f"{o.get('tax_type') or 'GST'} @ {o.get('tax_percent') or 0}%"
            tax_buckets[key]["count"] += 1
            tax_buckets[key]["tax_amount"] += o.get("tax_amount") or 0
            tax_buckets[key]["invoice_total"] += o.get("invoice_total") or 0
            tax_buckets[key]["revenue"] += o.get("operating_revenue") or 0
        else:
            non_tax_revenue += o.get("operating_revenue") or 0

    invoice = {
        "operating_revenue": revenue["total"],
        "tax_amount": sum(o.get("tax_amount") or 0 for o in orders),
        "invoice_total": sum(o.get("invoice_total") or 0 for o in orders),
        "non_taxable_revenue": non_tax_revenue,
        "by_tax_type": sorted(
            [{"label": k, **v} for k, v in tax_buckets.items()],
            key=lambda x: -x["invoice_total"],
        ),
    }

    # ---- Cost breakdown (6-way factory/outside + packing + freight) ----
    def item_sum(key):
        return sum((it.get(key) or 0) for o in orders for it in (o.get("items") or []))

    factory_complete = item_sum("factory_complete")
    factory_glass = item_sum("factory_glass")
    factory_fitting = item_sum("factory_fitting")
    outside_complete = item_sum("outside_complete")
    outside_glass = item_sum("outside_glass")
    outside_fitting = item_sum("outside_fitting")
    packing_cost_total = sum(o.get("packing_cost") or 0 for o in orders)
    freight_paid_total = sum(o.get("freight_paid") or 0 for o in orders)
    other_expense_total = sum(o.get("other_expense_total") or 0 for o in orders)

    # Other expenses grouped by description
    other_exp_by_desc = defaultdict(lambda: {"amount": 0, "count": 0})
    for o in orders:
        for e in (o.get("other_expense") or []):
            k = (e.get("description") or "Unlabelled").strip() or "Unlabelled"
            other_exp_by_desc[k]["amount"] += from_paise(to_paise(e.get("amount")))
            other_exp_by_desc[k]["count"] += 1

    cost = {
        "factory": {
            "total": factory_complete + factory_glass + factory_fitting,
            "complete": factory_complete,
            "glass": factory_glass,
            "fitting": factory_fitting,
        },
        "outside": {
            "total": outside_complete + outside_glass + outside_fitting,
            "complete": outside_complete,
            "glass": outside_glass,
            "fitting": outside_fitting,
        },
        "packing": packing_cost_total,
        "freight": freight_paid_total,
        "other_expense": other_expense_total,
        "other_expense_by_description": sorted(
            [{"description": k, **v} for k, v in other_exp_by_desc.items()],
            key=lambda x: -x["amount"],
        ),
        "total": (factory_complete + factory_glass + factory_fitting
                  + outside_complete + outside_glass + outside_fitting
                  + packing_cost_total + freight_paid_total + other_expense_total),
    }

    # ---- Profit by category ----
    prof_by_main = defaultdict(lambda: {"revenue": 0, "cost": 0, "profit": 0, "orders": 0})
    prof_sub_by_main = defaultdict(lambda: defaultdict(lambda: {"revenue": 0, "profit": 0}))
    for o in orders:
        rev = o.get("operating_revenue") or 0
        prof = o.get("net_profit") or 0
        item_sales = sum(i.get("product_sales") or 0 for i in o.get("items") or [])
        for it in o.get("items") or []:
            mc = it.get("main_category") or "Uncategorised"
            sc = it.get("sub_category") or "—"
            ratio = ((it.get("product_sales") or 0) / item_sales) if item_sales else 0
            r = rev * ratio
            p = prof * ratio
            prof_by_main[mc]["revenue"] += r
            prof_by_main[mc]["profit"] += p
            prof_by_main[mc]["cost"] += (r - p)
            prof_sub_by_main[mc][sc]["revenue"] += r
            prof_sub_by_main[mc][sc]["profit"] += p
            prof_by_main[mc]["orders"] += 1

    profit = {
        "operating_revenue": revenue["total"],
        "total_cost": cost["total"],
        "net_profit": revenue["total"] - cost["total"],
        "margin_percent": ((revenue["total"] - cost["total"]) / revenue["total"] * 100) if revenue["total"] else 0,
        "by_main_category": sorted(
            [{"main_category": k, **v,
              "margin_percent": (v["profit"] / v["revenue"] * 100) if v["revenue"] else 0}
             for k, v in prof_by_main.items()],
            key=lambda x: -x["profit"],
        ),
        "by_sub_category": {
            mc: sorted([{"sub_category": s, **v} for s, v in subs.items()],
                       key=lambda x: -x["profit"])
            for mc, subs in prof_sub_by_main.items()
        },
    }

    # ---- Receivable (unpaid / partial orders) ----
    # Bug fix (2026-07-22): the receivable KPI must show the REMAINING
    # order balance (invoice_total − allocated), clamped to zero on
    # over-payment. Preserves the by_status shape but uses each order's
    # dashboard-clamped outstanding paise for the amount buckets and the
    # per-client roll-ups. Single-sourced with /api/dashboard.kpis.
    by_status = {"Unpaid": {"count": 0, "amount": 0}, "Partial": {"count": 0, "amount": 0}, "Paid": {"count": 0, "amount": 0}}
    receivable_orders = []
    for o in orders:
        st = o.get("payment_status") or "Unpaid"
        by_status.setdefault(st, {"count": 0, "amount": 0})
        by_status[st]["count"] += 1
        if st == "Paid":
            # Paid orders contribute 0 to the receivable KPI by definition.
            # We still keep the count in the by_status roll-up for context.
            continue
        # Cancelled orders would be status='Cancelled' on lifecycle side,
        # not payment_status — but the domain helper already clamps them
        # for us, so an "Unpaid Cancelled" order still resolves to 0.
        out_p = order_dashboard_outstanding_paise(o)
        if out_p <= 0:
            # Over-paid or fully-recovered "Unpaid" order → skip from the
            # order list too (would show as ₹0 which is noise).
            continue
        amt_display = from_paise(out_p)
        by_status[st]["amount"] += amt_display
        receivable_orders.append({
            "id": o.get("id"),
            "client_name": o.get("client_name"),
            "shipped_date": o.get("shipped_date"),
            "invoice_total": o.get("invoice_total") or 0,
            "outstanding_balance": amt_display,   # NEW — see Bug fix note
            "payment_status": st,
        })
    receivable_orders.sort(key=lambda x: -(x["outstanding_balance"] or 0))

    # aggregate receivable by client — same dashboard-clamped amount as above.
    recv_by_client = defaultdict(lambda: {"amount": 0, "orders": 0})
    for o in orders:
        out_p = order_dashboard_outstanding_paise(o)
        if out_p <= 0:
            continue
        recv_by_client[o.get("client_name") or "Unknown"]["amount"] += from_paise(out_p)
        recv_by_client[o.get("client_name") or "Unknown"]["orders"] += 1

    receivable = {
        "total": by_status.get("Unpaid", {}).get("amount", 0) + by_status.get("Partial", {}).get("amount", 0),
        "by_status": [{"status": k, **v} for k, v in by_status.items() if k in ("Unpaid", "Partial", "Paid")],
        "by_client": sorted(
            [{"client": k, **v} for k, v in recv_by_client.items()],
            key=lambda x: -x["amount"],
        )[:20],
        "orders": receivable_orders[:50],
    }

    # ---- Payable (money paid out — grouped by party / mode) ----
    # P0: sourced from canonical modules (customer_payments, purchase_payments,
    # cash_book_entries) — never from legacy db.payments.
    # P3: transfers are profit-neutral — excluded here.
    #
    # Phase 6 · Slice 2: active-record filtering is now delegated to the
    # domain layer. Fetch all rows, then filter with `is_*_active` /
    # `is_cash_book_entry_canonical`.
    cust_pays = await db.customer_payments.find({}, {"_id": 0}).to_list(20000)
    purchase_pays = await db.purchase_payments.find({}, {"_id": 0}).to_list(20000)
    cb_entries_all = await db.cash_book_entries.find({}, {"_id": 0}).to_list(20000)

    payable_by_party = defaultdict(lambda: {"paid": 0, "received": 0, "net": 0})
    payable_by_mode = defaultdict(lambda: {"paid": 0, "received": 0})
    total_paid = 0.0
    total_received = 0.0

    def _bump(party: str, mode: str, r: float, pd: float):
        nonlocal total_paid, total_received
        total_paid += pd
        total_received += r
        party = party or "Unknown"
        payable_by_party[party]["paid"] += pd
        payable_by_party[party]["received"] += r
        payable_by_party[party]["net"] = (payable_by_party[party]["received"]
                                          - payable_by_party[party]["paid"])
        m = mode or "Other"
        payable_by_mode[m]["paid"] += pd
        payable_by_mode[m]["received"] += r

    for p in cust_pays:
        if not is_customer_payment_active(p):
            continue
        amt = from_paise(to_paise(p.get("amount")))
        _bump(p.get("customer_name") or "Customer", p.get("mode") or "", amt, 0.0)
    for p in purchase_pays:
        if not is_purchase_payment_active(p):
            continue
        amt = from_paise(to_paise(p.get("amount")))
        _bump(p.get("vendor_name") or "Vendor", p.get("mode") or "", 0.0, amt)
    for e in cb_entries_all:
        if not is_cash_book_entry_canonical(e):
            continue
        amt = from_paise(to_paise(e.get("amount")))
        party = e.get("party_name") or ("Transfer" if e.get("kind") == "transfer" else "Cash Book")
        mode = e.get("mode") or ""
        if e.get("kind") == "general_income":
            _bump(party, mode, amt, 0.0)
        elif e.get("kind") == "general_expense":
            _bump(party, mode, 0.0, amt)
        # transfers are profit-neutral — excluded from payable roll-ups

    payable = {
        "total_paid": total_paid,
        "total_received": total_received,
        "net_out": total_paid - total_received,
        "by_party": sorted(
            [{"party": k, **v} for k, v in payable_by_party.items()],
            key=lambda x: -x["paid"],
        )[:25],
        "by_mode": [{"mode": k, **v} for k, v in payable_by_mode.items()],
    }

    # ---- Boxes ----
    boxes_used = sum(o.get("boxes_used") or 0 for o in orders)
    boxes_shipped = sum(o.get("boxes_shipped") or 0 for o in orders)
    boxes_by_transporter = defaultdict(lambda: {"boxes_shipped": 0, "orders": 0, "freight_paid": 0})
    for o in orders:
        t = o.get("transporter") or "Not set"
        boxes_by_transporter[t]["boxes_shipped"] += o.get("boxes_shipped") or 0
        boxes_by_transporter[t]["orders"] += 1
        boxes_by_transporter[t]["freight_paid"] += o.get("freight_paid") or 0

    boxes = {
        "used": boxes_used,
        "shipped": boxes_shipped,
        "gap": boxes_used - boxes_shipped,
        "packing_cost": packing_cost_total,
        "avg_cost_per_box": (packing_cost_total / boxes_used) if boxes_used else 0,
        "by_transporter": sorted(
            [{"transporter": k, **v} for k, v in boxes_by_transporter.items()],
            key=lambda x: -x["boxes_shipped"],
        ),
    }

    # ---- Freight ----
    freight_by_transporter = defaultdict(lambda: {"charged": 0, "paid": 0, "orders": 0, "boxes": 0})
    for o in orders:
        t = o.get("transporter") or "Not set"
        freight_by_transporter[t]["charged"] += o.get("freight_charged") or 0
        freight_by_transporter[t]["paid"] += o.get("freight_paid") or 0
        freight_by_transporter[t]["orders"] += 1
        freight_by_transporter[t]["boxes"] += o.get("boxes_shipped") or 0

    freight = {
        "charged": freight_charged_total,
        "paid": freight_paid_total,
        "recovery_gap": freight_charged_total - freight_paid_total,
        "by_transporter": sorted(
            [{"transporter": k, **v,
              "gap": v["charged"] - v["paid"]}
             for k, v in freight_by_transporter.items()],
            key=lambda x: -x["paid"],
        ),
    }

    return {
        "revenue": revenue,
        "invoice": invoice,
        "profit": profit,
        "cost": cost,
        "receivable": receivable,
        "payable": payable,
        "boxes": boxes,
        "freight": freight,
    }


# ================================================================
# META
# ================================================================
DEFAULT_MAIN_CATEGORIES = [
    "Chandelier", "Hanging Light", "Wall Light", "Table Lamp",
    "Ceiling Light", "Floor Lamp", "Candle Stand", "Glass",
]
DEFAULT_MODES = ["RHUF", "ICICI", "UPI", "Cash", "Raks"]
DEFAULT_TRANSPORTERS = ["Delhivery", "DTDC", "Gati", "VRL", "By Road", "Rail", "Self"]
TAX_TYPES = ["None", "GST", "IGST", "CGST_SGST"]
PAYMENT_STATUSES = ["Unpaid", "Partial", "Paid"]


@api_router.get("/meta")
async def meta():
    orders = await db.orders.find({}, {"_id": 0}).to_list(10000)

    main_cats = set(DEFAULT_MAIN_CATEGORIES)
    sub_by_main = defaultdict(set)
    products_by_sub = defaultdict(set)
    transporters = set(DEFAULT_TRANSPORTERS)
    clients = set()

    for o in orders:
        if o.get("client_name"):
            clients.add(o["client_name"])
        if o.get("transporter"):
            transporters.add(o["transporter"])
        for it in o.get("items") or []:
            mc = it.get("main_category")
            sc = it.get("sub_category") or ""
            pn = it.get("product_name") or ""
            if mc:
                main_cats.add(mc)
                if sc:
                    sub_by_main[mc].add(sc)
                    if pn:
                        products_by_sub[f"{mc}/{sc}"].add(pn)

    # Parties / modes list — derived from canonical sources plus legacy for continuity
    cust_names = await db.customer_payments.distinct("customer_name")
    vend_names_from_pp = await db.purchase_payments.distinct("vendor_name")
    cb_parties = await db.cash_book_entries.distinct("party_name")
    legacy_parties = await db.payments.distinct("party")
    parties = sorted({p for p in
                      (list(cust_names) + list(vend_names_from_pp)
                       + list(cb_parties) + list(legacy_parties))
                      if p})

    cust_modes = await db.customer_payments.distinct("mode")
    purch_modes = await db.purchase_payments.distinct("mode")
    cb_modes = await db.cash_book_entries.distinct("mode")
    legacy_modes = await db.payments.distinct("mode")
    modes = sorted(set(DEFAULT_MODES) | {m for m in
                                          (list(cust_modes) + list(purch_modes)
                                           + list(cb_modes) + list(legacy_modes))
                                          if m})

    accounts = await db.accounts.find({"archived": {"$ne": True}}, {"_id": 0}).sort("name", 1).to_list(500)

    vendor_docs = await db.vendors.find({}, {"_id": 0}).sort("name", 1).to_list(2000)
    vendors = [v.get("name") for v in vendor_docs if v.get("name")]
    # Also include vendors that appear on purchases even if not saved
    purch = await db.purchases.find({}, {"_id": 0, "vendor_name": 1}).to_list(5000)
    for p in purch:
        if p.get("vendor_name") and p["vendor_name"] not in vendors:
            vendors.append(p["vendor_name"])

    return {
        "main_categories": sorted(main_cats),
        "sub_categories_by_main": {k: sorted(v) for k, v in sub_by_main.items()},
        "products_by_sub": {k: sorted(v) for k, v in products_by_sub.items()},
        "clients": sorted(clients),
        "transporters": sorted(transporters),
        "parties": parties,
        "modes": modes,
        "tax_types": TAX_TYPES,
        "payment_statuses": PAYMENT_STATUSES,
        "payment_modes": PAYMENT_MODES,
        "account_types": ACCOUNT_TYPES,
        "accounts": accounts,
        "vendors": sorted(set(vendors)),
    }


# ================================================================
# EXPORTS
# ================================================================
INR_FMT = '"₹"#,##0'
DATE_FMT = "dd-mmm-yyyy"


def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            v = cell.value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max_len + 3, 42)


def _style_header(ws, ncols):
    header_fill = PatternFill("solid", fgColor="F5F3EC")
    header_font = Font(name="Calibri", size=11, bold=True, color="2C2A29")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _parse_date(v):
    if not v:
        return None
    try:
        return datetime.fromisoformat(v.replace("Z", "")).date()
    except Exception:
        return None


def _csv_response(rows: List[dict], fieldnames: List[str], filename: str):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@api_router.get("/export/orders.xlsx")
async def export_orders_xlsx():
    orders = await db.orders.find({}, {"_id": 0}).sort("shipped_date", 1).to_list(10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    headers = ["Order Date", "Shipped Date", "Client", "Payment Status", "Items",
               "Product Sales", "Freight Charged", "Packing Charged", "Other Revenue", "Operating Revenue",
               "Factory Cost", "Outside Cost", "Packing Cost", "Freight Paid",
               "Other Expense", "Total Cost",
               "Net Profit", "Margin %",
               "Tax Type", "Tax %", "Tax Amount", "Tax Manual?", "Invoice Total",
               "Boxes Used", "Boxes Shipped", "Transporter", "LR/Tracking", "Notes"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for o in orders:
        ws.append([
            _parse_date(o.get("order_date")),
            _parse_date(o.get("shipped_date")),
            o.get("client_name"),
            o.get("payment_status"),
            len(o.get("items") or []),
            o.get("product_sales_total") or 0,
            o.get("freight_charged") or 0,
            o.get("packing_recovery") or 0,
            o.get("other_revenue_total") or 0,
            o.get("operating_revenue") or 0,
            o.get("factory_cost_total") or 0,
            o.get("outside_cost_total") or 0,
            o.get("packing_cost") or 0,
            o.get("freight_paid") or 0,
            o.get("other_expense_total") or 0,
            o.get("total_cost") or 0,
            o.get("net_profit") or 0,
            round(o.get("margin_percent") or 0, 2),
            o.get("tax_type") or "None",
            o.get("tax_percent") or 0,
            o.get("tax_amount") or 0,
            "Yes" if o.get("tax_amount_manual") else "Auto",
            o.get("invoice_total") or 0,
            o.get("boxes_used") or 0,
            o.get("boxes_shipped") or 0,
            o.get("transporter") or "",
            o.get("lr_number") or "",
            o.get("notes") or "",
        ])

    money_cols = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 21, 23]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].number_format = DATE_FMT
        row[1].number_format = DATE_FMT
        for i in money_cols:
            row[i - 1].number_format = INR_FMT
        row[17].number_format = '0.0"%"'  # margin (col 18)
        row[19].number_format = '0.0"%"'  # tax % (col 20)

    _autofit(ws)
    return _xlsx_response(wb, "orders.xlsx")


@api_router.get("/export/order-items.xlsx")
async def export_order_items_xlsx():
    """One row per line item, denormalised with order info — most detailed export."""
    orders = await db.orders.find({}, {"_id": 0}).sort("shipped_date", 1).to_list(10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Order Items"

    headers = ["Shipped Date", "Client", "Main Category", "Sub Category", "Product",
               "Qty", "Rate", "Product Sales",
               "Factory Complete", "Factory Glass", "Factory Fitting",
               "Outside Complete", "Outside Glass", "Outside Fitting"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for o in orders:
        for it in o.get("items") or []:
            ws.append([
                _parse_date(o.get("shipped_date")),
                o.get("client_name"),
                it.get("main_category"),
                it.get("sub_category") or "",
                it.get("product_name"),
                it.get("qty") or 0,
                it.get("rate") or 0,
                it.get("product_sales") or 0,
                it.get("factory_complete") or 0,
                it.get("factory_glass") or 0,
                it.get("factory_fitting") or 0,
                it.get("outside_complete") or 0,
                it.get("outside_glass") or 0,
                it.get("outside_fitting") or 0,
            ])

    money_cols = list(range(7, 15))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].number_format = DATE_FMT
        for i in money_cols:
            row[i - 1].number_format = INR_FMT
    _autofit(ws)
    return _xlsx_response(wb, "order-items.xlsx")


@api_router.get("/export/orders.csv")
async def export_orders_csv():
    orders = await db.orders.find({}, {"_id": 0}).sort("shipped_date", 1).to_list(10000)
    rows = []
    for o in orders:
        rows.append({
            "order_date": o.get("order_date"),
            "shipped_date": o.get("shipped_date"),
            "client_name": o.get("client_name"),
            "payment_status": o.get("payment_status"),
            "item_count": len(o.get("items") or []),
            "product_sales_total": o.get("product_sales_total") or 0,
            "freight_charged": o.get("freight_charged") or 0,
            "operating_revenue": o.get("operating_revenue") or 0,
            "factory_cost_total": o.get("factory_cost_total") or 0,
            "outside_cost_total": o.get("outside_cost_total") or 0,
            "packing_cost": o.get("packing_cost") or 0,
            "freight_paid": o.get("freight_paid") or 0,
            "total_cost": o.get("total_cost") or 0,
            "net_profit": o.get("net_profit") or 0,
            "tax_type": o.get("tax_type"),
            "tax_percent": o.get("tax_percent"),
            "tax_amount": o.get("tax_amount") or 0,
            "invoice_total": o.get("invoice_total") or 0,
            "boxes_used": o.get("boxes_used") or 0,
            "boxes_shipped": o.get("boxes_shipped") or 0,
            "transporter": o.get("transporter"),
            "lr_number": o.get("lr_number"),
            "notes": o.get("notes"),
        })
    fields = list(rows[0].keys()) if rows else ["order_date"]
    return _csv_response(rows, fields, "orders.csv")


async def _build_cashbook_export_rows() -> List[dict]:
    """Union of Cash Book timeline sources — used by CSV/XLSX exports.
    Adds a `source_module` column so audits can trace every rupee back to
    its originating module.
    """
    cust_pays = await db.customer_payments.find({}, {"_id": 0}).to_list(20000)
    purchase_pays = await db.purchase_payments.find({}, {"_id": 0}).to_list(20000)
    cb_entries = await db.cash_book_entries.find({}, {"_id": 0}).to_list(20000)
    legacy = await db.payments.find({}, {"_id": 0}).to_list(20000)

    rows: List[dict] = []
    for p in cust_pays:
        rows.append({
            "date": p.get("date"),
            "source_module": "Sales Payment",
            "kind": "customer_payment",
            "party": p.get("customer_name") or "",
            "mode": p.get("mode") or "",
            "account_name": p.get("account_name") or "",
            "received": float(p.get("amount") or 0),
            "paid": 0.0,
            "reference": p.get("reference") or "",
            "note": p.get("remarks") or "",
            "source_id": p.get("id"),
        })
    for p in purchase_pays:
        rows.append({
            "date": p.get("date"),
            "source_module": "Purchase Payment",
            "kind": "vendor_payment",
            "party": p.get("vendor_name") or "",
            "mode": p.get("mode") or "",
            "account_name": p.get("account_name") or "",
            "received": 0.0,
            "paid": float(p.get("amount") or 0),
            "reference": p.get("reference") or "",
            "note": p.get("remarks") or "",
            "source_id": p.get("id"),
        })
    for e in cb_entries:
        kind = e.get("kind") or ""
        received = float(e.get("amount") or 0) if kind == "general_income" else 0.0
        paid = float(e.get("amount") or 0) if kind == "general_expense" else 0.0
        if kind == "transfer":
            # emit two rows so both ledgers can be traced (net-zero on P&L)
            rows.append({
                "date": e.get("date"),
                "source_module": "Transfer",
                "kind": "transfer_out",
                "party": e.get("to_account_name") or "",
                "mode": e.get("mode") or "",
                "account_name": e.get("from_account_name") or "",
                "received": 0.0,
                "paid": float(e.get("amount") or 0),
                "reference": e.get("reference") or "",
                "note": e.get("notes") or "",
                "source_id": e.get("id"),
            })
            rows.append({
                "date": e.get("date"),
                "source_module": "Transfer",
                "kind": "transfer_in",
                "party": e.get("from_account_name") or "",
                "mode": e.get("mode") or "",
                "account_name": e.get("to_account_name") or "",
                "received": float(e.get("amount") or 0),
                "paid": 0.0,
                "reference": e.get("reference") or "",
                "note": e.get("notes") or "",
                "source_id": e.get("id"),
            })
            continue
        source = "Cash Book" if e.get("source") in (None, "cash_book") else "Cash Book (Legacy Shim)"
        rows.append({
            "date": e.get("date"),
            "source_module": source,
            "kind": kind or "general",
            "party": e.get("party_name") or "",
            "mode": e.get("mode") or "",
            "account_name": e.get("account_name") or "",
            "received": received,
            "paid": paid,
            "reference": e.get("reference") or "",
            "note": e.get("notes") or "",
            "source_id": e.get("id"),
        })
    for d in legacy:
        received = float((d.get("received_by_me") or 0) + (d.get("received_by_fac") or 0))
        paid = float((d.get("payment_by_me") or 0) + (d.get("payment_by_fac") or 0))
        rows.append({
            "date": d.get("date"),
            "source_module": "Migration",
            "kind": "legacy",
            "party": d.get("party") or "",
            "mode": d.get("mode") or "",
            "account_name": "",
            "received": received,
            "paid": paid,
            "reference": "",
            "note": d.get("note") or "",
            "source_id": d.get("id"),
        })
    rows.sort(key=lambda r: (r.get("date") or ""))
    return rows


@api_router.get("/export/payments.csv")
async def export_payments_csv():
    rows = await _build_cashbook_export_rows()
    fields = ["date", "source_module", "kind", "party", "mode", "account_name",
              "received", "paid", "reference", "note", "source_id"]
    return _csv_response(rows, fields, "cash-book.csv")


@api_router.get("/export/payments.xlsx")
async def export_payments_xlsx():
    rows = await _build_cashbook_export_rows()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Book"

    headers = ["Date", "Source Module", "Kind", "Party", "Mode", "Account",
               "Received", "Paid", "Net", "Reference", "Note", "Source ID"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in rows:
        received = float(r.get("received") or 0)
        paid = float(r.get("paid") or 0)
        ws.append([
            _parse_date(r.get("date")),
            r.get("source_module") or "",
            r.get("kind") or "",
            r.get("party") or "",
            r.get("mode") or "",
            r.get("account_name") or "",
            received,
            paid,
            received - paid,
            r.get("reference") or "",
            r.get("note") or "",
            r.get("source_id") or "",
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].number_format = DATE_FMT
        for i in (7, 8, 9):
            row[i - 1].number_format = INR_FMT
    _autofit(ws)
    return _xlsx_response(wb, "cash-book.xlsx")


# ================================================================
# PARTY LEDGER — single source of truth
# ================================================================
def _short_id(x: Optional[str]) -> str:
    return (x or "")[:8]


@api_router.get("/party-ledger/summary")
async def party_ledger_summary():
    """Return a list of every party with its current outstanding / advance / last txn.
    Sources:
      - orders.invoice_total → debit (billed)
      - customer_payments.amount → credit (paid)
      - customer_payments.unallocated → advance
    """
    orders = await db.orders.find({}, {"_id": 0}).to_list(20000)
    payments = await db.customer_payments.find({}, {"_id": 0}).to_list(20000)
    known = {c["name"] async for c in db.customers.find({}, {"_id": 0, "name": 1})}

    parties: dict = {}

    def _get(name):
        n = name or "Unknown"
        if n not in parties:
            parties[n] = {
                "party": n, "total_billed": 0.0, "total_received": 0.0,
                "allocated": 0.0, "advance": 0.0, "outstanding": 0.0,
                "orders": 0, "payments": 0, "last_txn_date": None,
            }
        return parties[n]

    for o in orders:
        p = _get(o.get("client_name"))
        p["total_billed"] += float(o.get("invoice_total") or 0)
        p["orders"] += 1
        d = o.get("last_shipped_date") or o.get("shipped_date") or o.get("order_date")
        if d and (p["last_txn_date"] is None or d > p["last_txn_date"]):
            p["last_txn_date"] = d

    for pay in payments:
        p = _get(pay.get("customer_name"))
        amt = float(pay.get("amount") or 0)
        adv = float(pay.get("unallocated") or 0)
        p["total_received"] += amt
        p["allocated"] += (amt - adv)
        p["advance"] += adv
        p["payments"] += 1
        d = pay.get("date")
        if d and (p["last_txn_date"] is None or d > p["last_txn_date"]):
            p["last_txn_date"] = d

    for name in known:
        _get(name)

    for p in parties.values():
        # outstanding = billed - allocated (advance is a separate positive credit)
        p["outstanding"] = round(p["total_billed"] - p["allocated"], 2)

    rows = sorted(
        parties.values(),
        key=lambda x: (-abs(x["outstanding"]), -(x["advance"] or 0), x["party"]),
    )
    return {
        "count": len(rows),
        "total_outstanding": round(sum(p["outstanding"] for p in rows if p["outstanding"] > 0), 2),
        "total_advance": round(sum(p["advance"] for p in rows), 2),
        "parties": rows,
    }


@api_router.get("/party-ledger")
async def party_ledger(party: str = Query(..., min_length=1)):
    """Chronological statement for one party.
    Events:
      - Invoice (Debit) — one per order (uses invoice_total, dated by last shipment date)
      - Payment (Credit) — from customer_payments; includes allocation details
    Also includes legacy cash-book payments (`payments` collection) for parity.
    """
    orders = await db.orders.find({"client_name": party}, {"_id": 0}).to_list(5000)
    cust_pays = await db.customer_payments.find({"customer_name": party}, {"_id": 0}).sort("date", 1).to_list(5000)
    legacy = await db.payments.find({"party": party}, {"_id": 0}).sort("date", 1).to_list(5000)

    order_by_id = {o.get("id"): o for o in orders}
    events: List[dict] = []

    for o in orders:
        inv = float(o.get("invoice_total") or 0)
        if inv <= 0 and not (o.get("shipments") or []):
            continue
        # one aggregate invoice line per order (uses stored invoice_total which already
        # reflects only shipped qty). Show # of shipments in meta.
        events.append({
            "type": "invoice",
            "date": o.get("last_shipped_date") or o.get("shipped_date") or o.get("order_date"),
            "ref": _short_id(o.get("id")),
            "description": f"Invoice · Order #{_short_id(o.get('id'))}",
            "debit": inv,
            "credit": 0.0,
            "order_id": o.get("id"),
            "meta": {
                "status": o.get("status"),
                "items": len(o.get("items") or []),
                "shipments": len(o.get("shipments") or []),
                "shipped_qty": o.get("shipped_qty_total") or 0,
                "ordered_qty": o.get("ordered_qty_total") or 0,
                "payment_status": o.get("payment_status"),
            },
        })

    for pay in cust_pays:
        amt = float(pay.get("amount") or 0)
        adv = float(pay.get("unallocated") or 0)
        alloc_list = pay.get("allocations") or []
        alloc_desc = []
        for a in alloc_list:
            oid = a.get("order_id")
            amt_a = float(a.get("amount") or 0)
            alloc_desc.append({
                "order_id": oid, "amount": amt_a,
                "short": _short_id(oid), "invoice_total": float((order_by_id.get(oid) or {}).get("invoice_total") or 0),
            })
        descr = f"Payment · {pay.get('mode') or 'Cash'}"
        if pay.get("account_name"):
            descr += f" → {pay.get('account_name')}"
        if pay.get("reference"):
            descr += f" · Ref {pay.get('reference')}"
        if adv > 0.5:
            descr += f" · Advance {adv:.0f}"
        events.append({
            "type": "payment",
            "date": pay.get("date"),
            "ref": pay.get("reference") or _short_id(pay.get("id")),
            "description": descr,
            "debit": 0.0,
            "credit": amt,
            "payment_id": pay.get("id"),
            "meta": {
                "mode": pay.get("mode"),
                "account_name": pay.get("account_name"),
                "account_id": pay.get("account_id"),
                "reference": pay.get("reference"),
                "remarks": pay.get("remarks"),
                "advance": adv,
                "allocations": alloc_desc,
            },
        })

    # Legacy cash-book entries (kept for parity so nothing is lost)
    for d in legacy:
        received = float((d.get("received_by_me") or 0) + (d.get("received_by_fac") or 0))
        paid = float((d.get("payment_by_me") or 0) + (d.get("payment_by_fac") or 0))
        if received == 0 and paid == 0:
            continue
        events.append({
            "type": "cashbook",
            "date": d.get("date"),
            "ref": _short_id(d.get("id")),
            "description": f"Cash Book · {d.get('mode') or 'Cash'}"
                           + (f" · {d.get('note')}" if d.get("note") else ""),
            "debit": paid,
            "credit": received,
            "cashbook_id": d.get("id"),
            "meta": {"mode": d.get("mode"), "note": d.get("note")},
        })

    # Sort chronologically; invoice before payment on the same date.
    def _k(e):
        return (e.get("date") or "", 0 if e["type"] == "invoice" else 1)
    events.sort(key=_k)

    balance = 0.0
    total_billed = 0.0
    total_received = 0.0
    for e in events:
        balance += float(e.get("debit") or 0) - float(e.get("credit") or 0)
        e["running_balance"] = round(balance, 2)
        total_billed += float(e.get("debit") or 0)
        total_received += float(e.get("credit") or 0)

    advance = sum(float(p.get("unallocated") or 0) for p in cust_pays)
    allocated = sum(float(p.get("amount") or 0) - float(p.get("unallocated") or 0) for p in cust_pays)
    outstanding = round(sum(float(o.get("invoice_total") or 0) for o in orders) - allocated, 2)

    return {
        "party": party,
        "count": len(events),
        "total_billed": round(total_billed, 2),
        "total_received": round(total_received, 2),
        "advance": round(advance, 2),
        "allocated": round(allocated, 2),
        "outstanding": outstanding,
        "net_balance": round(total_billed - total_received, 2),
        "entries": events,
        "orders": [{
            "id": o.get("id"),
            "date": o.get("last_shipped_date") or o.get("shipped_date") or o.get("order_date"),
            "invoice_total": float(o.get("invoice_total") or 0),
            "total_received": float(o.get("total_received") or 0),
            "outstanding_balance": float(o.get("outstanding_balance") or 0),
            "payment_status": o.get("payment_status"),
            "status": o.get("status"),
        } for o in sorted(orders,
                          key=lambda x: (x.get("last_shipped_date") or x.get("order_date") or ""))],
    }


# ================================================================
# MIGRATION: old transactions collection → orders collection
# ================================================================
async def _migrate_transactions_to_orders() -> int:
    """Group old flat transactions rows into orders by client+shipped_date, and
    materialise one Shipment covering all items so historical numbers stay intact."""
    tx = await db.transactions.find({}, {"_id": 0}).to_list(20000)
    if not tx:
        return 0

    groups = defaultdict(list)
    for t in tx:
        day = (t.get("shipped_date") or "")[:10] or "undated"
        key = (t.get("client_name") or "Unknown").strip() + "||" + day
        groups[key].append(t)

    orders_to_insert = []
    for key, rows in groups.items():
        client_name, day = key.split("||")
        shipped_date = rows[0].get("shipped_date")

        packing_cost = sum(float(r.get("packing") or 0) for r in rows)
        freight_paid = sum(float(r.get("freight") or 0) for r in rows)

        items = []
        ship_items = []
        for r in rows:
            it = OrderItem(
                main_category=r.get("category") or "Uncategorised",
                sub_category="",
                product_name=r.get("particulars") or "",
                qty=float(r.get("qty") or 0),
                rate=float(r.get("rate") or 0),
                product_sales=float(r.get("sales") or 0),
                factory_complete=float(r.get("factory_complete") or 0),
                factory_glass=float(r.get("factory_glass") or 0),
                factory_fitting=float(r.get("factory_fitting") or 0),
                outside_complete=float(r.get("outside_complete") or 0),
                outside_glass=float(r.get("outside_glass") or 0),
                outside_fitting=float(r.get("outside_fitting") or 0),
            ).model_dump()
            items.append(it)
            ship_items.append({"order_item_id": it["id"], "qty": it["qty"]})

        # Single shipment covering all items — freight/packing rolled up
        auto_shipment = Shipment(
            date=shipped_date,
            items=[ShipmentItem(**s).model_dump() for s in ship_items],
            freight_paid=freight_paid,
            boxes_shipped=0,
            remarks="Auto-created from legacy data during migration",
        ).model_dump()

        order = Order(
            client_name=client_name,
            order_date=shipped_date,
            shipped_date=shipped_date,
            status="Fully Shipped",
            payment_status="Paid",  # historical rows treated as paid
            items=items,
            shipments=[auto_shipment],
            packing_cost=packing_cost,
        ).model_dump()
        compute_order_aggregates(order)
        orders_to_insert.append(order)

    if orders_to_insert:
        await db.orders.insert_many(orders_to_insert)

    for name in {o["client_name"] for o in orders_to_insert if o.get("client_name")}:
        await db.customers.update_one(
            {"name": name},
            {"$setOnInsert": Customer(name=name).model_dump()},
            upsert=True,
        )
    return len(orders_to_insert)


async def _migrate_order_payments_to_customer_payments() -> int:
    """Move each embedded order_payment out into a CustomerPayment doc with a single allocation."""
    count = 0
    async for o in db.orders.find({"order_payments": {"$exists": True, "$ne": []}}):
        oid = o["id"]
        client = o.get("client_name") or "Unknown"
        for p in (o.get("order_payments") or []):
            amt = float(p.get("amount") or 0)
            cp = CustomerPayment(
                customer_name=client,
                date=p.get("date"),
                amount=amt,
                mode=p.get("mode") or "Cash",
                account_id=p.get("account_id") or "",
                account_name=p.get("account_name") or "",
                reference=p.get("reference") or "",
                remarks=p.get("remarks") or "",
                allocations=[PaymentAllocation(order_id=oid, amount=amt).model_dump()],
                allocated_total=amt,
                unallocated=0,
            ).model_dump()
            await db.customer_payments.insert_one(cp)
            count += 1
        # Remove the embedded field to complete the move
        await db.orders.update_one({"id": oid}, {"$unset": {"order_payments": ""}})
    return count


@api_router.post("/migrate")
async def migrate_endpoint(force: bool = False, confirm_wipe: Optional[str] = None):
    """Migrate legacy transactions → orders. `force=true` wipes ONLY orders
    (never touches payments/accounts/customers). Requires confirm_wipe="YES" for safety."""
    order_count = await db.orders.count_documents({})
    if order_count > 0 and not force:
        return {"status": "skipped", "orders": order_count}
    if force:
        if confirm_wipe != "YES":
            raise HTTPException(
                status_code=400,
                detail="Refusing to wipe orders. Pass confirm_wipe=YES to proceed.",
            )
        await db.orders.delete_many({})
    n = await _migrate_transactions_to_orders()
    return {"status": "migrated", "orders_created": n}


# ================================================================
# SEED (kept for fresh installs — highly guarded to prevent accidental wipes)
# ================================================================
@api_router.post("/seed")
async def seed(force: bool = False, confirm_wipe: Optional[str] = None):
    """One-shot seed of legacy transactions + payments + orders migration.
    Contract:
      - No args: skips silently if any data exists (idempotent for fresh installs).
      - force=true: refuses unless confirm_wipe='YES' is also passed.
        When force=true+confirm_wipe='YES', wipes transactions/orders/payments and reseeds.
    NEVER touches accounts or customers collections."""
    tx_count = await db.transactions.count_documents({})
    order_count = await db.orders.count_documents({})
    pay_count = await db.payments.count_documents({})
    if not force and (tx_count > 0 or order_count > 0):
        return {"status": "skipped", "orders": order_count,
                "raw_transactions": tx_count, "payments": pay_count}

    if force:
        if confirm_wipe != "YES":
            raise HTTPException(
                status_code=400,
                detail=("Refusing to wipe data. Pass confirm_wipe=YES with force=true. "
                        "This will delete ALL transactions, orders and payments."),
            )
        logger.warning(
            f"SEED force=true confirmed — wiping {tx_count} transactions, "
            f"{order_count} orders, {pay_count} payments before reseed."
        )
        await db.transactions.delete_many({})
        await db.orders.delete_many({})
        await db.payments.delete_many({})

    seed_path = ROOT_DIR / "seed" / "pl_seed.json"
    if not seed_path.exists():
        return {"status": "no_seed_file"}
    data = json.loads(seed_path.read_text())

    # Import raw rows into "transactions" (legacy) so migration groups them
    tx_docs = []
    for r in data.get("pl", []):
        cat = r.get("category")
        if not cat or cat in ("-",):
            continue
        tx_docs.append({
            "id": str(uuid.uuid4()),
            "category": cat,
            "particulars": r.get("particulars") or "",
            "client_name": r.get("client_name") or "",
            "qty": float(r.get("qty") or 0),
            "rate": float(r.get("rate") or 0),
            "sales": float(r.get("sales") or 0),
            "factory_complete": float(r.get("factory_complete") or 0),
            "factory_glass": float(r.get("factory_glass") or 0),
            "factory_fitting": float(r.get("factory_fitting") or 0),
            "outside_complete": float(r.get("outside_complete") or 0),
            "outside_glass": float(r.get("outside_glass") or 0),
            "outside_fitting": float(r.get("outside_fitting") or 0),
            "packing": float(r.get("packing") or 0),
            "freight": float(r.get("freight") or 0),
            "net_profit": float(r.get("net_profit") or 0),
            "shipped_date": r.get("shipped_date"),
        })
    if tx_docs:
        await db.transactions.insert_many(tx_docs)

    pay_docs = []
    for r in data.get("cash_flow", []):
        if not r.get("party") and not r.get("date"):
            continue
        p = Payment(
            date=r.get("date"),
            received_by_me=float(r.get("received_by_me") or 0),
            received_by_fac=float(r.get("received_by_fac") or 0),
            payment_by_me=float(r.get("payment_by_me") or 0),
            payment_by_fac=float(r.get("payment_by_fac") or 0),
            party=r.get("party") or "Unknown",
            mode=r.get("mode") or "Cash",
        ).model_dump()
        # P0: seeded legacy rows are stamped so the Cash Book timeline can
        # surface them as read-only "Migration" entries without counting them
        # in dashboard KPIs.
        p["source"] = "legacy_migrated"
        pay_docs.append(p)
    if pay_docs:
        await db.payments.insert_many(pay_docs)

    # Now migrate raw → orders
    n = await _migrate_transactions_to_orders()
    return {"status": "seeded", "raw_transactions": len(tx_docs), "orders": n, "payments": len(pay_docs)}


@api_router.get("/")
async def root():
    return {"message": "Artisan Ledger API — order-based"}


# ================================================================
# PHASE 6 — Auth (JWT + bcrypt) + Admin Data Management
# ================================================================

from fastapi import Request, Response, Depends            # noqa: E402
from fastapi.responses import FileResponse                # noqa: E402


class LoginIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    email: str
    password: str


class BootstrapIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    email: str
    password: str
    name: Optional[str] = ""


class ReauthIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    password: str


class ResetIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    scope: Literal["clear_transaction_data", "full_reset"]
    confirmation_phrase: str
    understand_checkbox: bool = False
    password: str
    keep_accounts: bool = True
    create_backup_first: bool = True


async def _admin_dep(request: Request):
    """FastAPI dependency wrapper — verifies admin JWT."""
    return await require_admin(request, db)


# ─── Auth endpoints ────────────────────────────────────────────────────────

@api_router.get("/auth/status")
async def auth_status():
    """Bootstrap discovery: returns whether an admin exists."""
    count = await db.users.count_documents({"role": "admin"})
    return {"has_admin": count > 0,
            "environment": current_environment(),
            "reset_enabled": is_reset_enabled()}


@api_router.post("/admin/bootstrap")
async def admin_bootstrap(payload: BootstrapIn, response: Response):
    """One-time endpoint — rejects once ≥1 admin exists."""
    if await db.users.count_documents({"role": "admin"}) > 0:
        raise HTTPException(400, "An admin already exists — bootstrap is closed.")
    if not payload.email or "@" not in payload.email:
        raise HTTPException(400, "Valid email required.")
    if not payload.password or len(payload.password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters.")
    # Validate JWT_SECRET FIRST — otherwise we'd create an orphan admin
    # row that the client can never receive a token for.
    from auth import get_jwt_secret
    get_jwt_secret()
    doc = new_user_doc(payload.email, payload.password, payload.name or "", role="admin")
    await db.users.insert_one(doc)
    access = create_access_token(doc["id"], doc["email"], doc["role"])
    refresh = create_refresh_token(doc["id"])
    set_auth_cookies(response, access, refresh)
    return {"user": user_public(doc), "access_token": access}


@api_router.post("/auth/login")
async def auth_login(payload: LoginIn, response: Response):
    email = (payload.email or "").strip().lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user or not verify_password(payload.password, user.get("password_hash") or ""):
        raise HTTPException(401, "Invalid email or password.")
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"last_login_at": datetime.now(timezone.utc).isoformat()}},
    )
    access = create_access_token(user["id"], user["email"], user.get("role") or "admin")
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)
    return {"user": user_public(user), "access_token": access}


@api_router.post("/auth/logout")
async def auth_logout(response: Response):
    clear_auth_cookies(response)
    return {"ok": True}


@api_router.get("/auth/me")
async def auth_me(request: Request):
    user = await get_current_user_from_db(request, db)
    return user_public(user)


@api_router.post("/auth/reauth")
async def auth_reauth(payload: ReauthIn, request: Request):
    """Re-verify the admin password before a sensitive action (e.g. reset)."""
    user = await get_current_user_from_db(request, db)
    if not verify_password(payload.password, user.get("password_hash") or ""):
        raise HTTPException(401, "Password incorrect.")
    return {"ok": True, "reverified_at": datetime.now(timezone.utc).isoformat()}


# ─── Data Management endpoints (all admin-gated) ───────────────────────────

@api_router.post("/admin/data-reset/preview")
async def admin_preview_reset(payload: dict, admin=Depends(_admin_dep)):
    scope = (payload or {}).get("scope") or "clear_transaction_data"
    keep_accounts = bool((payload or {}).get("keep_accounts", True))
    result = await preview_reset(db, scope, keep_accounts=keep_accounts)
    await log_audit(db, "data_reset_preview", admin, extra={"scope": scope,
                                                             "keep_accounts": keep_accounts})
    return result


@api_router.post("/admin/data-reset/execute")
async def admin_execute_reset(payload: ResetIn, request: Request, admin=Depends(_admin_dep)):
    if not is_reset_enabled():
        raise HTTPException(403,
                            "Data reset is disabled by server configuration "
                            "(ALLOW_ADMIN_DATA_RESET=false).")
    # 1. Re-verify admin password
    if not verify_password(payload.password, admin.get("password_hash") or ""):
        raise HTTPException(401, "Password incorrect.")
    # 2. Confirmation phrase check
    expected = ("CLEAR TRANSACTION DATA" if payload.scope == "clear_transaction_data"
                else "FULL RESET SAMRAT GLASS ERP")
    if current_environment() == "production":
        expected = f"{expected} {datetime.now(timezone.utc).date().isoformat()}"
    if (payload.confirmation_phrase or "").strip() != expected:
        raise HTTPException(400, f"Confirmation phrase mismatch. Expected {expected!r}.")
    # 3. Checkbox
    if not payload.understand_checkbox:
        raise HTTPException(400, "You must confirm you understand the action is irreversible.")

    # Phase 5 hook — snapshot reconciliation BEFORE reset so we know the
    # baseline health of the DB going into the destructive operation.
    try:
        pre_reconcile = summarize_reconcile(await run_reconcile(db))
    except Exception as ex:
        pre_reconcile = {"error": f"pre-reset reconcile failed: {ex}"}

    # 4. Backup (must succeed before deletion)
    backup_meta = None
    if payload.create_backup_first:
        try:
            backup_meta = await create_backup(
                db, created_by=admin.get("email") or admin.get("id"),
                note=f"Pre-{payload.scope} backup",
            )
        except Exception as e:
            raise HTTPException(500, f"Backup failed — reset aborted: {e}")

    # 5. Execute
    report = await execute_reset(
        db, scope=payload.scope, admin=admin,
        backup_id=(backup_meta or {}).get("id"),
        keep_accounts=payload.keep_accounts,
    )
    report["backup"] = backup_meta
    report["pre_reset_reconcile"] = pre_reconcile
    # Phase 5 hook — snapshot reconciliation AFTER the reset. Expected shape
    # depends on the scope, but at minimum system_fathers_firm must survive.
    try:
        report["post_reset_reconcile"] = summarize_reconcile(await run_reconcile(db))
    except Exception as ex:
        report["post_reset_reconcile"] = {"error": f"post-reset reconcile failed: {ex}"}
    return report


@api_router.post("/admin/backups")
async def admin_create_backup(payload: dict, admin=Depends(_admin_dep)):
    note = (payload or {}).get("note") or ""
    meta = await create_backup(db, created_by=admin.get("email") or admin.get("id"), note=note)
    await log_audit(db, "backup_create", admin, extra={"backup_id": meta["id"]})
    return meta


@api_router.get("/admin/backups")
async def admin_list_backups(admin=Depends(_admin_dep)):
    return await list_backups(db)


@api_router.get("/admin/backups/{bid}/download")
async def admin_download_backup(bid: str, admin=Depends(_admin_dep)):
    meta = await get_backup_meta(db, bid)
    path = meta.get("storage_location")
    if not path or not os.path.exists(path):
        raise HTTPException(404, "Backup file missing on disk.")
    return FileResponse(path, media_type="application/zip", filename=meta.get("filename") or f"{bid}.zip")


@api_router.delete("/admin/backups/{bid}")
async def admin_delete_backup(bid: str, admin=Depends(_admin_dep)):
    result = await delete_backup(db, bid)
    await log_audit(db, "backup_delete", admin, extra={"backup_id": bid})
    return result


@api_router.post("/admin/test-dataset/load")
async def admin_load_test_dataset(admin=Depends(_admin_dep)):
    result = await load_test_dataset(db, admin=admin)
    await log_audit(db, "test_dataset_load", admin, extra={"dataset_id": result["test_dataset_id"]})
    return result


@api_router.delete("/admin/test-dataset/{dataset_id}")
async def admin_remove_test_dataset(dataset_id: str, admin=Depends(_admin_dep)):
    result = await remove_test_dataset(db, dataset_id)
    await log_audit(db, "test_dataset_remove", admin, extra={"dataset_id": dataset_id})
    return result


@api_router.get("/admin/audit-logs")
async def admin_get_audit_logs(admin=Depends(_admin_dep), limit: int = 200):
    return await list_audit_logs(db, limit=limit)


# ─── Phase 5 (P2) — /api/reconcile ─────────────────────────────────────────

@api_router.get("/reconcile")
async def api_reconcile_report(admin=Depends(_admin_dep)):
    """Read-only integrity report. GUARANTEED zero writes.
    Runs every domain invariant and returns the full structured report."""
    return await run_reconcile(db)


@api_router.post("/reconcile/run")
async def api_reconcile_run(admin=Depends(_admin_dep)):
    """Run the reconciliation and write exactly ONE audit log row.
    If the audit write fails, the reconciliation report is still returned
    with an `audit_warning` field so the caller is aware."""
    report = await run_reconcile(db)
    try:
        await log_audit(db, "reconcile_run", admin, extra={
            "summary": summarize_reconcile(report),
        })
    except Exception as ex:
        report["audit_warning"] = f"Audit log write failed: {ex}"
    return report


@api_router.get("/admin/reconcile/last")
async def api_reconcile_last(admin=Depends(_admin_dep)):
    """Return the most recent reconcile_run audit summary, or null."""
    doc = await db.admin_audit_logs.find_one(
        {"kind": "reconcile_run"}, {"_id": 0}, sort=[("at", -1)]
    )
    return doc or {}


# ================================================================
# PHASE 2 — Party identity admin endpoints
# ================================================================
@api_router.get("/party-migration/last-report")
async def get_last_party_migration_report():
    doc = await db.admin_migration_reports.find_one(
        {"phase": "P1_party_identity"}, {"_id": 0}, sort=[("created_at", -1)]
    )
    if not doc:
        return {"status": "no_report"}
    return doc


@api_router.post("/party-migration/run")
async def run_party_migration_now():
    """Idempotent — re-run the canonical party migration. Safe to call
    any time after Phase 2 shipped."""
    report = await run_party_migration(db)
    doc = {
        "id": str(uuid.uuid4()),
        "phase": "P1_party_identity",
        "created_at": datetime.now(timezone.utc).isoformat(),
        **report,
    }
    await db.admin_migration_reports.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


# ─── Bug fix (2026-07-22) · Canonical vendor_party_id backfill ──────────────
#
# Every Purchase and PurchasePayment must carry `vendor_party_id` pointing
# at the canonical `db.parties` row for its vendor. Historically the
# auto-generated Order → Purchase linkage set only `vendor_name`; renaming
# a vendor or two similarly-named parties made the display fragile.
#
# This migration is idempotent — safe to re-run any number of times. It
# NEVER guesses on ambiguous matches. Unmatched / ambiguous rows are
# reported so they can be resolved manually via the party rename / merge
# admin flow.

async def _resolve_vendor_party_for_doc(doc: dict) -> tuple[str | None, str]:
    """Resolve the vendor party for a purchase-like doc using the
    documented resolution order. Returns (party_id, resolution).

    Resolution order (deterministic, non-destructive):
      1. Existing valid `vendor_party_id` on doc + party actually exists.
      2. Factory alias on `vendor_name` → SYSTEM_FF_ID.
      3. `linked_supplier_id == FACTORY_SUPPLIER_ID` → SYSTEM_FF_ID.
      4. Existing `linked_supplier_id` matches a vendor with `party_id`.
      5. Normalized `vendor_name` matches EXACTLY ONE non-system vendor
         party.
      6. Otherwise unmatched (may be created via `get_or_create_vendor_party`
         if the name is unambiguous).

    `resolution` is one of: "already_linked" · "factory_alias" ·
    "supplier_id_match" · "vendor_name_created" · "vendor_name_matched"
    · "ambiguous" · "unmatched".
    """
    # (1) already linked?
    existing_id = doc.get("vendor_party_id")
    if existing_id:
        exists = await db.parties.find_one({"id": existing_id}, {"_id": 1})
        if exists:
            return existing_id, "already_linked"

    vname = (doc.get("vendor_name") or "").strip()

    # (2) Factory / FF alias
    if is_ff_alias(vname):
        return SYSTEM_FF_ID, "factory_alias"

    # (3) linked_supplier_id === Factory
    supplier_id = doc.get("linked_supplier_id")
    if supplier_id and supplier_id == FACTORY_SUPPLIER_ID:
        return SYSTEM_FF_ID, "factory_alias"

    # (4) linked_supplier_id → vendors → party_id
    if supplier_id:
        v = await db.vendors.find_one(
            {"id": supplier_id}, {"_id": 0, "party_id": 1, "name": 1}
        )
        if v and v.get("party_id"):
            return v["party_id"], "supplier_id_match"

    # (5) Normalized vendor name → non-ambiguous vendor party
    if vname:
        # Ambiguity check: how many non-system, non-archived vendor parties
        # match this normalized name?
        from party_sync import normalize_name
        norm = normalize_name(vname)
        if norm:
            matches = await db.parties.find(
                {"type": "vendor", "normalized_name": norm,
                 "archived": {"$ne": True}, "is_system": {"$ne": True}},
                {"_id": 0, "id": 1, "name": 1},
            ).to_list(5)
            if len(matches) == 1:
                return matches[0]["id"], "vendor_name_matched"
            if len(matches) > 1:
                return None, "ambiguous"

        # Fall through — no exact match: create the canonical vendor party
        # via the shared helper. This mirrors how manual create_purchase
        # handled the missing-party case and stays consistent with
        # `run_party_migration` which already creates parties for
        # every distinct purchase vendor name.
        p = await get_or_create_vendor_party(db, vname)
        if p and p.get("id"):
            return p["id"], "vendor_name_created"

    return None, "unmatched"


async def _backfill_purchase_vendor_party_ids() -> dict:
    """Set `vendor_party_id` on every db.purchases + db.purchase_payments
    row per the resolution order above. Never guesses on ambiguous names.

    Report shape:
      {
        "purchases": {
          "scanned": int, "already_linked": int, "newly_linked": int,
          "ambiguous": [ids...], "unmatched": [ids...],
          "by_resolution": {"factory_alias": n, "supplier_id_match": n, ...},
        },
        "purchase_payments": { ...same shape... },
      }
    Idempotent — subsequent runs see previously linked rows as
    `already_linked` and take no writes for them.
    """
    def _blank():
        return {
            "scanned": 0, "already_linked": 0, "newly_linked": 0,
            "ambiguous": [], "unmatched": [],
            "by_resolution": {},
        }

    rep = {"purchases": _blank(), "purchase_payments": _blank()}

    async def _sweep(coll_name: str, section: str):
        bucket = rep[section]
        async for doc in db[coll_name].find({}, {"_id": 0}):
            bucket["scanned"] += 1
            pid, res = await _resolve_vendor_party_for_doc(doc)
            bucket["by_resolution"][res] = bucket["by_resolution"].get(res, 0) + 1
            if res == "already_linked":
                bucket["already_linked"] += 1
                continue
            if res == "ambiguous":
                bucket["ambiguous"].append({"id": doc.get("id"),
                                            "vendor_name": doc.get("vendor_name")})
                continue
            if res == "unmatched" or not pid:
                bucket["unmatched"].append({"id": doc.get("id"),
                                            "vendor_name": doc.get("vendor_name")})
                continue
            # Newly linked (factory_alias, supplier_id_match,
            # vendor_name_matched, vendor_name_created).
            await db[coll_name].update_one(
                {"id": doc["id"]}, {"$set": {"vendor_party_id": pid}}
            )
            bucket["newly_linked"] += 1

    await _sweep("purchases", "purchases")
    await _sweep("purchase_payments", "purchase_payments")
    return rep


@api_router.post("/admin/purchases/backfill-vendor-party-id")
async def admin_backfill_purchase_vendor_party_ids(
    admin=Depends(_admin_dep),
):
    """Idempotent — re-runs the vendor_party_id backfill on every
    db.purchases + db.purchase_payments row. Returns a structured
    migration report with counts + lists of ambiguous / unmatched rows
    for manual review."""
    report = await _backfill_purchase_vendor_party_ids()
    doc = {
        "id": str(uuid.uuid4()),
        "phase": "bug_fix_vendor_party_id_backfill",
        "created_at": datetime.now(timezone.utc).isoformat(),
        **report,
    }
    await db.admin_migration_reports.insert_one(doc)
    return {k: v for k, v in doc.items() if k != "_id"}


class PartyRenameIn(BaseModel):
    model_config = ConfigDict(extra="ignore")
    # Preferred field is `display_name`. `new_name` is accepted as an alias
    # for callers that document the older key.
    display_name: Optional[str] = None
    new_name: Optional[str] = None


@api_router.post("/parties/{pid}/rename")
async def party_rename(pid: str, payload: PartyRenameIn):
    """Rename a party in place — preserves the party_id, pushes the old
    normalized name onto `aliases`. This is a RENAME, not a reassignment;
    historical transactions continue to resolve through the same party_id.

    Accepts `display_name` (preferred) or the legacy alias `new_name`.
    """
    if pid == SYSTEM_FF_ID:
        raise HTTPException(400, "System Father's Firm party cannot be renamed.")
    name = (payload.display_name or payload.new_name or "").strip()
    if not name:
        raise HTTPException(400, "display_name (or new_name) is required.")
    p = await rename_party(db, pid, name)
    if not p:
        raise HTTPException(404, "Party not found")
    return p


# ================================================================
# PHASE 3 — First-Class Transfers (canonical source of truth)
# ================================================================
@api_router.post("/transfers", response_model=Transfer)
async def api_create_transfer(payload: TransferIn):
    return await create_transfer(db, payload)


@api_router.get("/transfers", response_model=List[Transfer])
async def api_list_transfers(
    kind: Optional[str] = None,
    account_id: Optional[str] = None,
    party_id: Optional[str] = None,
    include_reversed: bool = True,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 500,
):
    q: dict = {}
    if not include_reversed:
        q["status"] = "active"
    if kind:
        q["kind"] = kind
    if account_id:
        q["$or"] = [{"from_side.account_id": account_id},
                    {"to_side.account_id": account_id}]
    if party_id:
        q.setdefault("$or", [])
        q["$or"] += [{"from_side.party_id": party_id},
                     {"to_side.party_id": party_id}]
    if start_date or end_date:
        d: dict = {}
        if start_date: d["$gte"] = start_date
        if end_date: d["$lte"] = end_date
        q["date"] = d
    docs = await db.transfers.find(q, {"_id": 0}).sort("date", -1).to_list(limit)
    return [Transfer(**d) for d in docs]


@api_router.get("/transfers/{tid}", response_model=Transfer)
async def api_get_transfer(tid: str):
    d = await db.transfers.find_one({"id": tid}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Transfer not found")
    return Transfer(**d)


@api_router.put("/transfers/{tid}", response_model=Transfer)
async def api_edit_transfer(tid: str, payload: TransferIn):
    return await edit_transfer(db, tid, payload)


@api_router.post("/transfers/{tid}/reverse", response_model=Transfer)
async def api_reverse_transfer(tid: str):
    return await reverse_transfer(db, tid)


@api_router.delete("/transfers/{tid}", response_model=Transfer)
async def api_delete_transfer(tid: str):
    """Alias for /reverse — deletion is reversal, never a hard delete."""
    return await reverse_transfer(db, tid)


@api_router.get("/accounts/{aid}/balance")
async def api_account_balance(aid: str):
    return await derive_account_balance(db, aid)


@api_router.post("/transfer-migration/run")
async def api_run_transfer_migration():
    """Idempotent re-run — safe to call at any time. Reports counts."""
    report = await run_transfer_migration(db)
    report["ran_at"] = datetime.now(timezone.utc).isoformat()
    await db.admin_migration_reports.insert_one({
        "id": str(uuid.uuid4()),
        "created_at": datetime.now(timezone.utc).isoformat(),
        **report,
    })
    return report


# ================================================================
# CASH BOOK — canonical write path for general income/expense/transfer
# and unified timeline read view.
# ================================================================

def _validate_cbe(cbe: dict) -> None:
    kind = cbe.get("kind")
    amt = float(cbe.get("amount") or 0)
    if amt <= 0:
        raise HTTPException(400, "Amount must be greater than zero.")
    if kind == "transfer":
        f = (cbe.get("from_account_id") or cbe.get("from_account_name") or "").strip()
        t = (cbe.get("to_account_id") or cbe.get("to_account_name") or "").strip()
        if not f or not t:
            raise HTTPException(400, "Transfer requires both from- and to-accounts.")
        if f == t and cbe.get("from_account_name", "") == cbe.get("to_account_name", ""):
            raise HTTPException(400, "Transfer source and destination must differ.")
    elif kind in ("general_income", "general_expense"):
        # party_name is optional; account is not strictly required but recommended.
        pass
    else:
        raise HTTPException(400, f"Unsupported Cash Book kind: {kind!r}")


@api_router.post("/cash-book-entries", response_model=CashBookEntry)
async def create_cash_book_entry(payload: CashBookEntryBase):
    cbe = CashBookEntry(**payload.model_dump()).model_dump()
    cbe["source"] = "cash_book"
    _validate_cbe(cbe)
    # Phase 3: transfers must be recorded in `db.transfers` (sole source of
    # truth). Auto-forward here for UI back-compat and return a synthetic
    # CashBookEntry shape pointing at the new canonical transfer.
    if cbe.get("kind") == "transfer":
        payload = TransferIn(
            date=cbe.get("date") or datetime.now(timezone.utc).isoformat()[:10],
            from_side=TransferSide(type="account",
                                   account_id=cbe.get("from_account_id"),
                                   account_name=cbe.get("from_account_name") or ""),
            to_side=TransferSide(type="account",
                                 account_id=cbe.get("to_account_id"),
                                 account_name=cbe.get("to_account_name") or ""),
            amount=float(cbe.get("amount") or 0),
            mode=cbe.get("mode") or "Bank Transfer",
            reference=cbe.get("reference") or "",
            notes=cbe.get("notes") or "",
        )
        t = await create_transfer(db, payload)
        cbe["id"] = t.id
        cbe["migrated_to_transfer_id"] = t.id
        # Do NOT insert into db.cash_book_entries — the transfer is the
        # canonical row; we still return a CashBookEntry shape so old
        # callers stay working.
        return CashBookEntry(**cbe)
    await db.cash_book_entries.insert_one(cbe)
    return CashBookEntry(**cbe)


@api_router.get("/cash-book-entries", response_model=List[CashBookEntry])
async def list_cash_book_entries(
    kind: Optional[str] = None,
    include_shim: bool = False,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    q: dict = {}
    if not include_shim:
        q["source"] = {"$ne": "legacy_shim"}
    if kind and kind != "all":
        q["kind"] = kind
    if start_date or end_date:
        d: dict = {}
        if start_date:
            d["$gte"] = start_date
        if end_date:
            d["$lte"] = end_date
        q["date"] = d
    docs = await db.cash_book_entries.find(q, {"_id": 0}).sort("date", -1).to_list(5000)
    return [CashBookEntry(**d) for d in docs]


@api_router.get("/cash-book-entries/{eid}", response_model=CashBookEntry)
async def get_cash_book_entry(eid: str):
    d = await db.cash_book_entries.find_one({"id": eid}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Cash Book entry not found")
    return CashBookEntry(**d)


@api_router.put("/cash-book-entries/{eid}", response_model=CashBookEntry)
async def update_cash_book_entry(eid: str, payload: CashBookEntryBase):
    existing = await db.cash_book_entries.find_one({"id": eid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Cash Book entry not found")
    if existing.get("source") == "legacy_shim":
        raise HTTPException(400, "Legacy shim rows cannot be edited via Cash Book.")
    cbe = {**existing, **payload.model_dump()}
    cbe["source"] = existing.get("source") or "cash_book"
    cbe["updated_at"] = datetime.now(timezone.utc).isoformat()
    _validate_cbe(cbe)
    await db.cash_book_entries.update_one({"id": eid}, {"$set": cbe})
    return CashBookEntry(**cbe)


@api_router.delete("/cash-book-entries/{eid}")
async def delete_cash_book_entry(eid: str):
    existing = await db.cash_book_entries.find_one({"id": eid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Cash Book entry not found")
    if existing.get("source") == "legacy_shim":
        raise HTTPException(400, "Legacy shim rows cannot be deleted via Cash Book.")
    await db.cash_book_entries.delete_one({"id": eid})
    return {"deleted": True}


# ─── Unified timeline (Cash Book UI feed) ──────────────────────────────────
# Every money-move in the ERP, projected into a common shape so the Cash Book
# page can render it as a chronological feed with source-linked navigation.

def _cb_row_from_customer_payment(p: dict) -> dict:
    amt = float(p.get("amount") or 0)
    return {
        "event_id": p.get("id"),
        "date": p.get("date"),
        "kind": "customer_payment",
        "source_module": "Sales Payments",
        "source_route": "/sales-payments",
        "title": "Customer Payment",
        "party": p.get("customer_name") or "",
        "mode": p.get("mode") or "",
        "account_name": p.get("account_name") or "",
        "received": amt,
        "paid": 0.0,
        "amount": amt,
        "direction": "in",
        "reference": p.get("reference") or "",
        "notes": p.get("remarks") or "",
        "source_document": {
            "collection": "customer_payments",
            "id": p.get("id"),
            "route": "/sales-payments",
            "linked_order_ids": [a.get("order_id") for a in (p.get("allocations") or [])
                                 if a.get("order_id")],
        },
        "editable": True,
    }


def _cb_row_from_purchase_payment(p: dict) -> dict:
    amt = float(p.get("amount") or 0)
    return {
        "event_id": p.get("id"),
        "date": p.get("date"),
        "kind": "vendor_payment",
        "source_module": "Purchase Payments",
        "source_route": "/purchase-payments",
        "title": "Vendor Payment",
        "party": p.get("vendor_name") or "",
        "mode": p.get("mode") or "",
        "account_name": p.get("account_name") or "",
        "received": 0.0,
        "paid": amt,
        "amount": amt,
        "direction": "out",
        "reference": p.get("reference") or "",
        "notes": p.get("remarks") or "",
        "source_document": {
            "collection": "purchase_payments",
            "id": p.get("id"),
            "route": "/purchase-payments",
            "linked_purchase_ids": [a.get("purchase_id") for a in (p.get("allocations") or [])
                                    if a.get("purchase_id")],
        },
        "editable": True,
    }


def _cb_row_from_cbe(e: dict) -> dict:
    amt = float(e.get("amount") or 0)
    kind = e.get("kind") or "general_expense"
    source_map = {
        "cash_book": "Cash Book",
        "legacy_shim": "Cash Book (Legacy Shim)",
        "legacy_migrated": "Migration",
    }
    source_module = source_map.get(e.get("source") or "cash_book", "Cash Book")
    title_map = {
        "general_income": "General Income",
        "general_expense": "General Expense",
        "transfer": "Transfer",
    }
    if kind == "transfer":
        received = amt
        paid = amt
        direction = "transfer"
    elif kind == "general_income":
        received = amt
        paid = 0.0
        direction = "in"
    else:
        received = 0.0
        paid = amt
        direction = "out"
    return {
        "event_id": e.get("id"),
        "date": e.get("date"),
        "kind": kind,
        "source_module": source_module,
        "source_route": "/payments",
        "title": title_map.get(kind, "Cash Book"),
        "party": e.get("party_name") or "",
        "mode": e.get("mode") or "",
        "account_name": e.get("account_name") or "",
        "from_account_name": e.get("from_account_name") or "",
        "to_account_name": e.get("to_account_name") or "",
        "received": received,
        "paid": paid,
        "amount": amt,
        "direction": direction,
        "reference": e.get("reference") or "",
        "notes": e.get("notes") or "",
        "source_document": {
            "collection": "cash_book_entries",
            "id": e.get("id"),
            "route": "/payments",
        },
        "editable": e.get("source") == "cash_book",
    }


def _cb_row_from_transfer(t: dict) -> dict:
    """Phase 3: canonical transfer → one Cash Book timeline row.
    Never emits two duplicate rows — this is the whole point of moving
    to `db.transfers` as the single source of truth."""
    fs = t.get("from_side") or {}
    ts = t.get("to_side") or {}
    def _label(side: dict) -> str:
        if side.get("type") == "party":
            return side.get("party_name") or "Father's Firm"
        return side.get("account_name") or ""
    return {
        "event_id": t.get("id"),
        "date": t.get("date"),
        "kind": "transfer",
        "source_module": "Transfer",
        "source_route": "/transfers",
        "title": {"account_to_account": "Transfer",
                  "rakshit_to_ff": "Transfer to Father's Firm",
                  "ff_to_rakshit": "Transfer from Father's Firm"}.get(t.get("kind"), "Transfer"),
        "party": "",
        "mode": t.get("mode") or "",
        "account_name": "",
        "from_account_name": _label(fs),
        "to_account_name": _label(ts),
        "received": 0.0,
        "paid": 0.0,
        "amount": float(t.get("amount") or 0),
        "direction": "transfer",
        "reference": t.get("reference") or "",
        "notes": t.get("notes") or "",
        "source_document": {"collection": "transfers", "id": t.get("id"),
                            "route": "/transfers"},
        "editable": True,
    }


def _cb_row_from_legacy(p: dict) -> dict:
    received = float((p.get("received_by_me") or 0) + (p.get("received_by_fac") or 0))
    paid = float((p.get("payment_by_me") or 0) + (p.get("payment_by_fac") or 0))
    return {
        "event_id": p.get("id"),
        "date": p.get("date"),
        "kind": "legacy",
        "source_module": "Migration",
        "source_route": "/payments",
        "title": "Legacy Cash Book",
        "party": p.get("party") or "",
        "mode": p.get("mode") or "",
        "account_name": "",
        "received": received,
        "paid": paid,
        "amount": received if received >= paid else paid,
        "direction": "in" if received > paid else ("out" if paid > 0 else "flat"),
        "reference": "",
        "notes": p.get("note") or "",
        "source_document": {
            "collection": "payments",
            "id": p.get("id"),
            "route": "/payments",
        },
        "editable": False,
    }


@api_router.get("/cash-book")
async def cash_book_timeline(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    source_module: Optional[str] = None,
    kind: Optional[str] = None,
    party: Optional[str] = None,
    include_shim: bool = False,
    include_migration: bool = True,
    limit: int = 500,
):
    """Unified Cash Book timeline — one chronological feed sourced from every
    canonical financial module. Rows are read-only projections; edits must be
    performed on the origin document (see `source_document`).
    """
    cust = await db.customer_payments.find({}, {"_id": 0}).sort("date", -1).to_list(5000)
    purch = await db.purchase_payments.find({}, {"_id": 0}).sort("date", -1).to_list(5000)
    # Cash Book entries — but SUPPRESS any transfer-kind CBE that has been
    # migrated to db.transfers (Phase 3 deterministic migration marker).
    cbe_q: dict = {"$or": [
        {"kind": {"$ne": "transfer"}},
        {"kind": "transfer", "migrated_to_transfer_id": {"$exists": False}},
    ]}
    if not include_shim:
        cbe_q["source"] = {"$ne": "legacy_shim"}
    cbe = await db.cash_book_entries.find(cbe_q, {"_id": 0}).sort("date", -1).to_list(5000)
    # Phase 3: transfers are the canonical timeline row for every transfer.
    transfers = await db.transfers.find(
        {"status": "active", "reverses_transfer_id": None},
        {"_id": 0},
    ).sort("date", -1).to_list(5000)
    legacy = []
    if include_migration:
        legacy = await db.payments.find({}, {"_id": 0}).sort("date", -1).to_list(5000)

    rows = ([_cb_row_from_customer_payment(p) for p in cust]
            + [_cb_row_from_purchase_payment(p) for p in purch]
            + [_cb_row_from_cbe(e) for e in cbe]
            + [_cb_row_from_transfer(t) for t in transfers]
            + [_cb_row_from_legacy(p) for p in legacy])

    def _pass(r: dict) -> bool:
        if start_date and (r.get("date") or "") < start_date:
            return False
        if end_date and (r.get("date") or "") > end_date:
            return False
        if source_module and (r.get("source_module") or "") != source_module:
            return False
        if kind and (r.get("kind") or "") != kind:
            return False
        if party and party.lower() not in (r.get("party") or "").lower():
            return False
        return True

    rows = [r for r in rows if _pass(r)]
    rows.sort(key=lambda r: (r.get("date") or ""), reverse=True)

    total_received = sum(float(r.get("received") or 0) for r in rows
                         if r.get("kind") in ("customer_payment", "general_income"))
    total_paid = sum(float(r.get("paid") or 0) for r in rows
                     if r.get("kind") in ("vendor_payment", "general_expense"))

    return {
        "count": len(rows),
        "total_received": round(total_received, 2),
        "total_paid": round(total_paid, 2),
        "net": round(total_received - total_paid, 2),
        "rows": rows[:limit],
    }


# ─── Business Events (unified ERP activity feed) ────────────────────────────

@api_router.get("/business-events")
async def business_events(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    event_type: Optional[str] = None,
    source_module: Optional[str] = None,
    limit: int = 500,
):
    """Firehose of every domain event across the ERP — Orders, Shipments,
    Purchases, Sales/Purchase Payments, Cash Book. Consumed by dashboards,
    activity feeds, notifications, audit trails. Every row conforms to a
    common Business Event envelope."""
    events: List[dict] = []

    async for o in db.orders.find({}, {"_id": 0}):
        events.append({
            "event_id": f"order_created:{o.get('id')}",
            "event_type": "order_created",
            "source_module": "Orders",
            "source_document": {"collection": "orders", "id": o.get("id"),
                                "route": f"/orders"},
            "date": o.get("order_date") or (o.get("created_at") or "")[:10],
            "created_at": o.get("created_at"),
            "updated_at": o.get("updated_at"),
            "party": o.get("client_name") or "",
            "amount": float(o.get("invoice_total") or 0),
            "title": "Order Created",
            "notes": o.get("notes") or "",
            "reversed": False,
            "created_by": "user",
        })
        for i, sh in enumerate(sorted((o.get("shipments") or []),
                                      key=lambda s: s.get("date") or "")):
            qty = sum(float((si or {}).get("qty") or 0) for si in (sh.get("items") or []))
            events.append({
                "event_id": f"shipment:{sh.get('id')}",
                "event_type": "shipment",
                "source_module": "Orders",
                "source_document": {"collection": "orders", "id": o.get("id"),
                                    "shipment_id": sh.get("id"), "route": "/orders"},
                "date": sh.get("date"),
                "created_at": sh.get("date"),
                "updated_at": sh.get("date"),
                "party": o.get("client_name") or "",
                "amount": float(sh.get("freight_paid") or 0),
                "title": f"Shipment {i + 1}",
                "notes": f"{qty:.0f} pcs · {sh.get('transporter') or ''}",
                "reversed": False,
                "created_by": "user",
            })

    async for p in db.customer_payments.find({}, {"_id": 0}):
        events.append({
            "event_id": f"customer_payment:{p.get('id')}",
            "event_type": "customer_payment",
            "source_module": "Sales Payments",
            "source_document": {"collection": "customer_payments", "id": p.get("id"),
                                "route": "/sales-payments"},
            "date": p.get("date"),
            "created_at": p.get("created_at"),
            "updated_at": p.get("created_at"),
            "party": p.get("customer_name") or "",
            "amount": float(p.get("amount") or 0),
            "title": "Customer Payment",
            "notes": p.get("remarks") or "",
            "reversed": False,
            "created_by": "user",
        })

    async for p in db.purchase_payments.find({}, {"_id": 0}):
        events.append({
            "event_id": f"vendor_payment:{p.get('id')}",
            "event_type": "vendor_payment",
            "source_module": "Purchase Payments",
            "source_document": {"collection": "purchase_payments", "id": p.get("id"),
                                "route": "/purchase-payments"},
            "date": p.get("date"),
            "created_at": p.get("created_at"),
            "updated_at": p.get("created_at"),
            "party": p.get("vendor_name") or "",
            "amount": float(p.get("amount") or 0),
            "title": "Vendor Payment",
            "notes": p.get("remarks") or "",
            "reversed": False,
            "created_by": "user",
        })

    async for pu in db.purchases.find({}, {"_id": 0}):
        events.append({
            "event_id": f"purchase_created:{pu.get('id')}",
            "event_type": "purchase_created",
            "source_module": "Purchases",
            "source_document": {"collection": "purchases", "id": pu.get("id"),
                                "route": "/purchases"},
            "date": pu.get("purchase_date"),
            "created_at": pu.get("created_at"),
            "updated_at": pu.get("updated_at"),
            "party": pu.get("vendor_name") or "",
            "amount": float(pu.get("invoice_total") or 0),
            "title": "Purchase Created",
            "notes": pu.get("notes") or "",
            "reversed": False,
            "created_by": "user",
        })

    async for e in db.cash_book_entries.find({"source": {"$ne": "legacy_shim"}}, {"_id": 0}):
        et = e.get("kind") or "general_expense"
        title_map = {"general_income": "General Income",
                     "general_expense": "General Expense",
                     "transfer": "Transfer"}
        events.append({
            "event_id": f"{et}:{e.get('id')}",
            "event_type": et,
            "source_module": "Cash Book",
            "source_document": {"collection": "cash_book_entries", "id": e.get("id"),
                                "route": "/payments"},
            "date": e.get("date"),
            "created_at": e.get("created_at"),
            "updated_at": e.get("updated_at"),
            "party": e.get("party_name") or "",
            "amount": float(e.get("amount") or 0),
            "title": title_map.get(et, "Cash Book"),
            "notes": e.get("notes") or "",
            "reversed": bool(e.get("reversed")),
            "created_by": "user",
        })

    def _pass(ev):
        if start_date and (ev.get("date") or "") < start_date:
            return False
        if end_date and (ev.get("date") or "") > end_date:
            return False
        if event_type and ev.get("event_type") != event_type:
            return False
        if source_module and ev.get("source_module") != source_module:
            return False
        return True

    events = [e for e in events if _pass(e)]
    events.sort(key=lambda e: (e.get("date") or "", e.get("created_at") or ""), reverse=True)
    return {"count": len(events), "events": events[:limit]}


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    # Phase 6: httpOnly auth cookies require an EXPLICIT origin (browsers reject
    # wildcard '*' with credentials). CORS_ORIGINS may hold an explicit
    # comma-separated allow-list; when unset we rely purely on the regex below,
    # which whitelists every preview.emergentagent.com subdomain + localhost.
    allow_origins=[o.strip() for o in (os.environ.get("CORS_ORIGINS") or "").split(",") if o.strip() and o.strip() != "*"],
    allow_origin_regex=r"https?://.*\.preview\.emergentagent\.com|https?://localhost(:\d+)?|https?://127\.0\.0\.1(:\d+)?",
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


async def _refresh_stored_aggregates() -> int:
    """Recompute and $set aggregate fields on every order using current formula.
    Idempotent — safe to run on startup. Fixes drift when the formula changed.

    Bug fix (2026-07-22): also refreshes the derived `shipped_date` and
    `last_shipped_date` on every order so historical fully-shipped orders
    with blank stored `shipped_date` get backfilled from their shipment
    records automatically on the next backend restart.
    """
    n = 0
    async for doc in db.orders.find({}):
        doc.pop("_id", None)
        before = {k: doc.get(k) for k in ("operating_revenue", "total_cost", "net_profit",
                                          "invoice_total", "estimated_operating_revenue",
                                          "estimated_net_profit", "shipped_date",
                                          "last_shipped_date")}
        compute_order_aggregates(doc)
        after = {k: doc.get(k) for k in ("operating_revenue", "total_cost", "net_profit",
                                         "invoice_total", "estimated_operating_revenue",
                                         "estimated_net_profit", "shipped_date",
                                         "last_shipped_date")}
        if before != after or "total_received" not in doc or "estimated_operating_revenue" not in before:
            await db.orders.update_one({"id": doc["id"]}, {"$set": {
                "product_sales_total": doc["product_sales_total"],
                "factory_cost_total": doc["factory_cost_total"],
                "outside_cost_total": doc["outside_cost_total"],
                "other_revenue_total": doc.get("other_revenue_total", 0),
                "other_expense_total": doc.get("other_expense_total", 0),
                "operating_revenue": doc["operating_revenue"],
                "total_cost": doc["total_cost"],
                "tax_amount": doc["tax_amount"],
                "invoice_total": doc["invoice_total"],
                "net_profit": doc["net_profit"],
                "margin_percent": doc["margin_percent"],
                "total_received": doc.get("total_received", 0),
                "outstanding_balance": doc.get("outstanding_balance", 0),
                # Phase 4 — estimated + realized aliases
                "estimated_factory_cost_total": doc["estimated_factory_cost_total"],
                "estimated_outside_cost_total": doc["estimated_outside_cost_total"],
                "estimated_operating_revenue": doc["estimated_operating_revenue"],
                "estimated_total_cost": doc["estimated_total_cost"],
                "estimated_net_profit": doc["estimated_net_profit"],
                "estimated_margin_percent": doc["estimated_margin_percent"],
                "realized_revenue": doc["realized_revenue"],
                "realized_net_profit": doc["realized_net_profit"],
                "revenue_recognized": doc["revenue_recognized"],
                "unrealized_revenue": doc["unrealized_revenue"],
                "unrealized_net_profit": doc["unrealized_net_profit"],
                # Bug fix (2026-07-22): derived completion date
                "shipped_date": doc.get("shipped_date"),
                "last_shipped_date": doc.get("last_shipped_date"),
            }})
            n += 1
    return n


# ================================================================
# SHIPMENTS (per-order)
# ================================================================
@api_router.post("/orders/{oid}/shipments", response_model=Order)
async def add_shipment(oid: str, payload: Shipment):
    order = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    order["shipments"] = (order.get("shipments") or []) + [payload.model_dump()]
    compute_order_aggregates(order)
    order["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.orders.update_one({"id": oid}, {"$set": order})
    await _recompute_payment_aggregates_for_orders([oid])
    updated = await db.orders.find_one({"id": oid}, {"_id": 0})
    return Order(**updated)


@api_router.put("/orders/{oid}/shipments/{sid}", response_model=Order)
async def update_shipment(oid: str, sid: str, payload: Shipment):
    order = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    shipments = order.get("shipments") or []
    idx = next((i for i, s in enumerate(shipments) if s.get("id") == sid), -1)
    if idx == -1:
        raise HTTPException(404, "Shipment not found")
    new_ship = payload.model_dump()
    new_ship["id"] = sid  # preserve
    shipments[idx] = new_ship
    order["shipments"] = shipments
    compute_order_aggregates(order)
    order["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.orders.update_one({"id": oid}, {"$set": order})
    await _recompute_payment_aggregates_for_orders([oid])
    updated = await db.orders.find_one({"id": oid}, {"_id": 0})
    return Order(**updated)


@api_router.delete("/orders/{oid}/shipments/{sid}")
async def delete_shipment(oid: str, sid: str):
    order = await db.orders.find_one({"id": oid}, {"_id": 0})
    if not order:
        raise HTTPException(404, "Order not found")
    before = len(order.get("shipments") or [])
    order["shipments"] = [s for s in (order.get("shipments") or []) if s.get("id") != sid]
    if len(order["shipments"]) == before:
        raise HTTPException(404, "Shipment not found")
    compute_order_aggregates(order)
    await db.orders.update_one({"id": oid}, {"$set": order})
    await _recompute_payment_aggregates_for_orders([oid])
    return {"deleted": True}


# ================================================================
# CUSTOMER PAYMENTS  (payments + multi-order allocation + advances)
# ================================================================
def _finalise_customer_payment(cp: dict) -> dict:
    amt = float(cp.get("amount") or 0)
    alloc = sum(float((a or {}).get("amount") or 0) for a in (cp.get("allocations") or []))
    if alloc > amt + 0.01:
        raise HTTPException(400, "Allocated amount exceeds payment amount")
    cp["allocated_total"] = alloc
    cp["unallocated"] = max(0.0, amt - alloc)
    return cp


@api_router.post("/customer-payments", response_model=CustomerPayment)
async def create_customer_payment(payload: CustomerPaymentBase):
    cp = CustomerPayment(**payload.model_dump()).model_dump()
    _finalise_customer_payment(cp)
    # Phase 2: canonical customer party id
    party = await get_or_create_customer_party(db, cp.get("customer_name") or "")
    if party:
        cp["customer_party_id"] = party["id"]
    await db.customer_payments.insert_one(cp)
    order_ids = [a.get("order_id") for a in (cp.get("allocations") or []) if a.get("order_id")]
    await _recompute_payment_aggregates_for_orders(order_ids)
    return CustomerPayment(**cp)


@api_router.get("/customer-payments", response_model=List[CustomerPayment])
async def list_customer_payments(
    customer_name: Optional[str] = None,
    account_id: Optional[str] = None,
    mode: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    only_with_advance: bool = False,
):
    q: dict = {}
    if customer_name:
        q["customer_name"] = {"$regex": customer_name, "$options": "i"}
    if account_id:
        q["account_id"] = account_id
    if mode and mode != "all":
        q["mode"] = mode
    if start_date or end_date:
        d: dict = {}
        if start_date: d["$gte"] = start_date
        if end_date: d["$lte"] = end_date
        q["date"] = d
    if only_with_advance:
        q["unallocated"] = {"$gt": 0}
    docs = await db.customer_payments.find(q, {"_id": 0}).sort("date", -1).to_list(5000)
    return [CustomerPayment(**d) for d in docs]


@api_router.get("/customer-payments/{pid}", response_model=CustomerPayment)
async def get_customer_payment(pid: str):
    doc = await db.customer_payments.find_one({"id": pid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Payment not found")
    return CustomerPayment(**doc)


@api_router.put("/customer-payments/{pid}", response_model=CustomerPayment)
async def update_customer_payment(pid: str, payload: CustomerPaymentBase):
    existing = await db.customer_payments.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Payment not found")
    old_order_ids = [a.get("order_id") for a in (existing.get("allocations") or []) if a.get("order_id")]
    cp = {**existing, **payload.model_dump()}
    _finalise_customer_payment(cp)
    party = await get_or_create_customer_party(db, cp.get("customer_name") or "")
    if party:
        cp["customer_party_id"] = party["id"]
    await db.customer_payments.update_one({"id": pid}, {"$set": cp})
    new_order_ids = [a.get("order_id") for a in (cp.get("allocations") or []) if a.get("order_id")]
    await _recompute_payment_aggregates_for_orders(list(set(old_order_ids + new_order_ids)))
    return CustomerPayment(**cp)


@api_router.delete("/customer-payments/{pid}")
async def delete_customer_payment(pid: str):
    existing = await db.customer_payments.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Payment not found")
    order_ids = [a.get("order_id") for a in (existing.get("allocations") or []) if a.get("order_id")]
    await db.customer_payments.delete_one({"id": pid})
    await _recompute_payment_aggregates_for_orders(order_ids)
    return {"deleted": True}


# ================================================================
# ACCOUNTS MASTER
# ================================================================
@api_router.get("/accounts", response_model=List[Account])
async def list_accounts(include_archived: bool = False):
    q = {} if include_archived else {"archived": {"$ne": True}}
    docs = await db.accounts.find(q, {"_id": 0}).sort("name", 1).to_list(1000)
    return [Account(**d) for d in docs]


@api_router.post("/accounts", response_model=Account)
async def create_account(payload: Account):
    data = payload.model_dump()
    data["id"] = str(uuid.uuid4())
    data["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.accounts.insert_one(data)
    return Account(**data)


@api_router.put("/accounts/{aid}", response_model=Account)
async def update_account(aid: str, payload: Account):
    existing = await db.accounts.find_one({"id": aid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Account not found")
    upd = payload.model_dump()
    upd.pop("id", None)
    upd.pop("created_at", None)
    await db.accounts.update_one({"id": aid}, {"$set": upd})
    doc = await db.accounts.find_one({"id": aid}, {"_id": 0})
    # Update denormalised account_name inside orders where this account is referenced
    if upd.get("name") and upd["name"] != existing.get("name"):
        await db.orders.update_many(
            {"order_payments.account_id": aid},
            {"$set": {"order_payments.$[p].account_name": upd["name"]}},
            array_filters=[{"p.account_id": aid}],
        )
    return Account(**doc)


@api_router.post("/accounts/{aid}/archive")
async def archive_account(aid: str, archived: bool = True):
    res = await db.accounts.update_one({"id": aid}, {"$set": {"archived": archived}})
    if res.matched_count == 0:
        raise HTTPException(404, "Account not found")
    return {"archived": archived}


async def _seed_default_accounts_if_empty():
    n = await db.accounts.count_documents({})
    if n > 0:
        return 0
    defaults = [
        ("ICICI Current", "Bank"),
        ("HDFC Current", "Bank"),
        ("Cash", "Cash"),
        ("Petty Cash", "PettyCash"),
        ("PhonePe", "UPI"),
        ("Google Pay", "UPI"),
        ("Paytm", "UPI"),
    ]
    for name, typ in defaults:
        await db.accounts.insert_one(Account(name=name, type=typ).model_dump())
    return len(defaults)


# ================================================================
# SALES PAYMENTS REPORT — sourced from customer_payments collection
# ================================================================
@api_router.get("/sales-payments")
async def sales_payments_report(
    account_id: Optional[str] = None,
    mode: Optional[str] = None,
    client_name: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    q: dict = {}
    if account_id:
        q["account_id"] = account_id
    if mode and mode != "all":
        q["mode"] = mode
    if client_name:
        q["customer_name"] = {"$regex": client_name, "$options": "i"}
    if start_date or end_date:
        d: dict = {}
        if start_date: d["$gte"] = start_date
        if end_date: d["$lte"] = end_date
        q["date"] = d

    docs = await db.customer_payments.find(q, {"_id": 0}).sort("date", -1).to_list(20000)
    rows = []
    total = 0
    total_advance = 0
    by_account = defaultdict(lambda: {"amount": 0, "count": 0})
    by_mode = defaultdict(lambda: {"amount": 0, "count": 0})
    for p in docs:
        amt = float(p.get("amount") or 0)
        adv = float(p.get("unallocated") or 0)
        total += amt
        total_advance += adv
        row = {
            "id": p.get("id"),
            "customer_name": p.get("customer_name"),
            "date": p.get("date"),
            "amount": amt,
            "mode": p.get("mode"),
            "account_id": p.get("account_id"),
            "account_name": p.get("account_name"),
            "reference": p.get("reference"),
            "remarks": p.get("remarks"),
            "allocated": float(p.get("allocated_total") or 0),
            "advance": adv,
            "allocations": p.get("allocations") or [],
        }
        rows.append(row)
        ak = row["account_name"] or "Unassigned"
        by_account[ak]["amount"] += amt
        by_account[ak]["count"] += 1
        mk = row["mode"] or "—"
        by_mode[mk]["amount"] += amt
        by_mode[mk]["count"] += 1

    return {
        "count": len(rows),
        "total": total,
        "total_advance": total_advance,
        "payments": rows,
        "by_account": sorted([{"account": k, **v} for k, v in by_account.items()],
                             key=lambda x: -x["amount"]),
        "by_mode": sorted([{"mode": k, **v} for k, v in by_mode.items()],
                          key=lambda x: -x["amount"]),
    }


@api_router.post("/orders/refresh-aggregates")
async def refresh_aggregates_endpoint():
    n = await _refresh_stored_aggregates()
    return {"refreshed": n}


# ================================================================
# VENDORS + PURCHASES + PURCHASE PAYMENTS (Accounts Payable module)
# ================================================================
class Vendor(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    gstin: Optional[str] = ""
    address: Optional[str] = ""
    notes: Optional[str] = ""
    archived: bool = False
    # Phase 2: cross-link to canonical parties row
    party_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class PurchaseItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category: Optional[str] = ""   # optional grouping (Raw Material / Glass / Fittings / Services…)
    description: str = ""
    qty: float = 0
    rate: float = 0
    amount: float = 0  # server-computed = qty * rate (if amount not provided)


class PurchaseBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    vendor_name: str
    purchase_date: Optional[str] = None
    invoice_no: Optional[str] = ""
    items: List[PurchaseItem] = []
    freight: float = 0
    other_charges: float = 0
    tax_applicable: bool = False
    tax_type: TaxType = "None"
    tax_percent: float = 0
    tax_amount: float = 0
    tax_amount_manual: bool = False
    notes: Optional[str] = ""
    payment_status: PaymentStatus = "Unpaid"


class Purchase(PurchaseBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    subtotal: float = 0
    invoice_total: float = 0
    total_paid: float = 0
    outstanding_balance: float = 0
    # Phase 2: stable vendor party id
    vendor_party_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    # Auto-linked-from-order metadata (present only when source_type =
    # 'order_product_purchase'). Manual purchases leave these blank / null.
    source_type: Optional[str] = None
    linked_to_order_id: Optional[str] = None
    linked_source_key: Optional[str] = None
    linked_supplier_id: Optional[str] = None
    linked_source_row_id: Optional[str] = None
    linked_order_item_id: Optional[str] = None
    linked_cost_category: Optional[str] = None
    stale: bool = False


class PurchasePaymentAllocation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    purchase_id: str
    amount: float = 0


class PurchasePaymentBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    vendor_name: str
    date: Optional[str] = None
    amount: float
    mode: str = "Cash"
    account_id: Optional[str] = None
    account_name: Optional[str] = ""
    reference: Optional[str] = ""
    remarks: Optional[str] = ""
    allocations: List[PurchasePaymentAllocation] = []
    # Party-ledger linkage. If paid_by_party_id is set (non-self), a linked
    # effect is posted on that party (e.g. Father's Firm paid on Rakshit's
    # behalf). split_paid_by_amount lets the payment be shared between
    # Rakshit and paid_by_party — the primary vendor reduction is the full
    # amount, while paid_by_party gets a linked entry for only its share.
    paid_by_party_id: Optional[str] = None
    paid_by_party_name: Optional[str] = None
    split_paid_by_amount: Optional[float] = None


class PurchasePayment(PurchasePaymentBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    allocated_total: float = 0
    unallocated: float = 0  # vendor advance
    # Phase 2: stable vendor party id
    vendor_party_id: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


def compute_purchase(purchase: dict) -> dict:
    """Recompute stored aggregates on a purchase (vendor bill).

    Phase 6 · Slice 4: this function is now a THIN ADAPTER over the
    pure domain helper `purchase_realized_amounts`. The item-mutation
    (stamping `it["amount"] = qty*rate` when amount was missing) stays
    here — it's a document-write concern, not pure calculation.
    """
    items = purchase.get("items") or []
    for it in items:
        if not it.get("amount"):
            it["amount"] = float(it.get("qty") or 0) * float(it.get("rate") or 0)
    real = purchase_realized_amounts(purchase)
    purchase["subtotal"] = from_paise(real["subtotal_paise"])
    purchase["tax_amount"] = from_paise(real["tax_amount_paise"])
    purchase["invoice_total"] = from_paise(real["invoice_total_paise"])
    purchase["items"] = items
    return purchase


async def _recompute_purchase_payment_aggregates(purchase_ids: List[str] = None):
    """Recompute total_paid, outstanding_balance and payment_status for given purchases.

    Phase 6 · Slice 4: allocation sums and outstanding derivation routed
    through `sum_allocations_to_purchase`. Paise-safe, idempotent.
    """
    q = {"id": {"$in": purchase_ids}} if purchase_ids else {}
    purchases = await db.purchases.find(q, {"_id": 0}).to_list(20000)
    for p in purchases:
        pid = p["id"]
        pays = await db.purchase_payments.find(
            {"allocations.purchase_id": pid}, {"_id": 0}
        ).to_list(10000)
        alloc_p = sum_allocations_to_purchase(pays, pid)
        invoice_p = to_paise(p.get("invoice_total"))
        # Pre-refactor rule: PURCHASE outstanding is CLAMPED to zero on
        # over-payment (unlike customer orders, which store the negative).
        # Preserve that exact asymmetry — matches `purchase_outstanding_from_alloc`.
        outstanding_p = purchase_outstanding_from_alloc(p, alloc_p)
        # 50-paise (₹0.50) hysteresis matches the pre-refactor rule.
        if alloc_p <= 50:
            status = "Unpaid"
        elif alloc_p + 50 >= invoice_p:
            status = "Paid"
        else:
            status = "Partial"
        await db.purchases.update_one(
            {"id": pid},
            {"$set": {
                "total_paid": from_paise(alloc_p),
                "outstanding_balance": from_paise(outstanding_p),
                "payment_status": status,
            }},
        )


# ---- Vendors CRUD ----
@api_router.get("/vendors", response_model=List[Vendor])
async def list_vendors(archived: bool = False):
    q = {} if archived else {"archived": {"$ne": True}}
    docs = await db.vendors.find(q, {"_id": 0}).sort("name", 1).to_list(2000)
    return [Vendor(**d) for d in docs]


@api_router.get("/purchase-sources")
async def list_purchase_sources():
    """Return the full set of pickable purchase sources for the Order dialog —
    Factory (Father's Firm — protected) as the first entry, followed by all
    active outside vendors. Kept small on purpose so the dropdown can be built
    from a single fetch."""
    docs = await db.vendors.find({"archived": {"$ne": True}}, {"_id": 0}).sort("name", 1).to_list(5000)
    rows = [{
        "id": FACTORY_SUPPLIER_ID,
        "name": "Factory",
        "type": "factory",
        "protected": True,
        "hint": "Father's Firm — settled via the Father's Firm ledger",
    }]
    for v in docs:
        rows.append({
            "id": v.get("id"),
            "name": v.get("name"),
            "type": "vendor",
            "protected": False,
            "hint": v.get("phone") or "",
        })
    return {"count": len(rows), "sources": rows}


@api_router.post("/vendors", response_model=Vendor)
async def create_vendor(payload: Vendor):
    exists = await db.vendors.find_one({"name": payload.name})
    if exists:
        raise HTTPException(400, "Vendor with this name already exists")
    v = payload.model_dump()
    # Phase 2: mirror into canonical parties + cross-link party_id.
    if is_ff_alias(v.get("name")):
        raise HTTPException(400, "'Factory' / 'Father's Firm' is a protected system party — cannot be created as a vendor.")
    party = await sync_vendor_directory(db, v.get("name") or "", vendor_id=v.get("id"),
                                        phone=v.get("phone") or "", gstin=v.get("gstin") or "")
    if party:
        v["party_id"] = party["id"]
    await db.vendors.insert_one(v)
    return Vendor(**v)


@api_router.put("/vendors/{vid}", response_model=Vendor)
async def update_vendor(vid: str, payload: Vendor):
    existing = await db.vendors.find_one({"id": vid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Vendor not found")
    data = payload.model_dump()
    data["id"] = vid
    # Phase 2: rename via stable party_id — never mutate history rows.
    pid = existing.get("party_id")
    if pid and data.get("name") and data["name"].strip() != (existing.get("name") or "").strip():
        await rename_party(db, pid, data["name"])
    else:
        # First time we see this vendor — ensure a canonical party exists.
        party = await sync_vendor_directory(db, data.get("name") or "", vendor_id=vid,
                                            phone=data.get("phone") or "",
                                            gstin=data.get("gstin") or "")
        if party:
            data["party_id"] = party["id"]
    res = await db.vendors.update_one({"id": vid}, {"$set": data})
    if res.matched_count == 0:
        raise HTTPException(404, "Vendor not found")
    return Vendor(**data)


@api_router.delete("/vendors/{vid}")
async def delete_vendor(vid: str):
    res = await db.vendors.delete_one({"id": vid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Vendor not found")
    return {"deleted": True}


# ---- Purchases CRUD ----
@api_router.get("/purchases", response_model=List[Purchase])
async def list_purchases(
    vendor_name: Optional[str] = None,
    payment_status: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    q: dict = {}
    if vendor_name:
        q["vendor_name"] = {"$regex": vendor_name, "$options": "i"}
    if payment_status and payment_status != "all":
        q["payment_status"] = payment_status
    if start_date or end_date:
        d: dict = {}
        if start_date: d["$gte"] = start_date
        if end_date: d["$lte"] = end_date
        q["purchase_date"] = d
    docs = await db.purchases.find(q, {"_id": 0}).sort("purchase_date", -1).to_list(5000)
    return [Purchase(**d) for d in docs]


@api_router.post("/purchases", response_model=Purchase)
async def create_purchase(payload: PurchaseBase):
    data = payload.model_dump()
    purchase = Purchase(**data).model_dump()
    compute_purchase(purchase)
    purchase["updated_at"] = datetime.now(timezone.utc).isoformat()
    # Phase 2: canonical vendor party id (Factory→system_fathers_firm).
    vname = data.get("vendor_name") or ""
    party = None
    if vname:
        party = (await resolve_party(db, ptype="vendor", display_name=vname)
                 if is_ff_alias(vname)
                 else await get_or_create_vendor_party(db, vname))
    if party:
        purchase["vendor_party_id"] = party["id"]
    await db.purchases.insert_one(purchase)
    # Ensure vendor exists in vendors master
    if data.get("vendor_name") and not is_ff_alias(data["vendor_name"]):
        await db.vendors.update_one(
            {"name": data["vendor_name"]},
            {"$setOnInsert": Vendor(name=data["vendor_name"]).model_dump()},
            upsert=True,
        )
    await _recompute_purchase_payment_aggregates([purchase["id"]])
    fresh = await db.purchases.find_one({"id": purchase["id"]}, {"_id": 0})
    return Purchase(**fresh)


@api_router.get("/purchases/{pid}", response_model=Purchase)
async def get_purchase(pid: str):
    d = await db.purchases.find_one({"id": pid}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Purchase not found")
    return Purchase(**d)


@api_router.put("/purchases/{pid}", response_model=Purchase)
async def update_purchase(pid: str, payload: PurchaseBase):
    existing = await db.purchases.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Purchase not found")
    data = payload.model_dump()
    data["id"] = pid
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    compute_purchase(data)
    await db.purchases.update_one({"id": pid}, {"$set": data})
    await _recompute_purchase_payment_aggregates([pid])
    fresh = await db.purchases.find_one({"id": pid}, {"_id": 0})
    return Purchase(**fresh)


@api_router.delete("/purchases/{pid}")
async def delete_purchase(pid: str):
    res = await db.purchases.delete_one({"id": pid})
    if res.deleted_count == 0:
        raise HTTPException(404, "Purchase not found")
    # Remove allocations referring to this purchase from purchase_payments
    async for pay in db.purchase_payments.find({"allocations.purchase_id": pid}, {"_id": 0}):
        new_alloc = [a for a in (pay.get("allocations") or []) if a.get("purchase_id") != pid]
        alloc_total = sum(float(a.get("amount") or 0) for a in new_alloc)
        await db.purchase_payments.update_one(
            {"id": pay["id"]},
            {"$set": {
                "allocations": new_alloc,
                "allocated_total": alloc_total,
                "unallocated": max(0.0, float(pay.get("amount") or 0) - alloc_total),
            }},
        )
    return {"deleted": True}


@api_router.get("/vendors/{name}/outstanding-purchases")
async def vendor_outstanding_purchases(name: str):
    """Return purchases for vendor that are Unpaid/Partial (for allocation UI).

    Phase 6 · Slice 4: monetary reads routed through to_paise/from_paise
    for paise-safe rounding at the response boundary.
    """
    docs = await db.purchases.find(
        {"vendor_name": name, "payment_status": {"$in": ["Unpaid", "Partial"]}},
        {"_id": 0},
    ).sort("purchase_date", 1).to_list(500)
    return [{
        "id": d["id"],
        "invoice_no": d.get("invoice_no") or "—",
        "purchase_date": d.get("purchase_date"),
        "invoice_total": from_paise(to_paise(d.get("invoice_total"))),
        "total_paid": from_paise(to_paise(d.get("total_paid"))),
        "outstanding_balance": from_paise(to_paise(d.get("outstanding_balance"))),
    } for d in docs]


# ---- Purchase Payments CRUD ----
@api_router.get("/purchase-payments", response_model=List[PurchasePayment])
async def list_purchase_payments(
    vendor_name: Optional[str] = None,
    mode: Optional[str] = None,
    account_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    only_with_advance: bool = False,
):
    q: dict = {}
    if vendor_name:
        q["vendor_name"] = {"$regex": vendor_name, "$options": "i"}
    if mode and mode != "all":
        q["mode"] = mode
    if account_id:
        q["account_id"] = account_id
    if only_with_advance:
        q["unallocated"] = {"$gt": 0}
    if start_date or end_date:
        d: dict = {}
        if start_date: d["$gte"] = start_date
        if end_date: d["$lte"] = end_date
        q["date"] = d
    docs = await db.purchase_payments.find(q, {"_id": 0}).sort("date", -1).to_list(20000)
    return [PurchasePayment(**d) for d in docs]


@api_router.post("/purchase-payments", response_model=PurchasePayment)
async def create_purchase_payment(payload: PurchasePaymentBase):
    p = PurchasePayment(**payload.model_dump()).model_dump()
    amt = float(p.get("amount") or 0)
    alloc = sum(float(a.get("amount") or 0) for a in (p.get("allocations") or []))
    p["allocated_total"] = round(alloc, 2)
    p["unallocated"] = round(max(0.0, amt - alloc), 2)
    # Phase 2: canonical vendor party id (Factory→system_fathers_firm).
    vname = p.get("vendor_name") or ""
    party = None
    if vname:
        party = (await resolve_party(db, ptype="vendor", display_name=vname)
                 if is_ff_alias(vname)
                 else await get_or_create_vendor_party(db, vname))
    if party:
        p["vendor_party_id"] = party["id"]
    await db.purchase_payments.insert_one(p)
    # Ensure vendor exists (skip Factory / FF alias)
    if p.get("vendor_name") and not is_ff_alias(p["vendor_name"]):
        await db.vendors.update_one(
            {"name": p["vendor_name"]},
            {"$setOnInsert": Vendor(name=p["vendor_name"]).model_dump()},
            upsert=True,
        )
    pids = [a["purchase_id"] for a in (p.get("allocations") or []) if a.get("purchase_id")]
    if pids:
        await _recompute_purchase_payment_aggregates(pids)
    return PurchasePayment(**p)


@api_router.get("/purchase-payments/{pid}", response_model=PurchasePayment)
async def get_purchase_payment(pid: str):
    d = await db.purchase_payments.find_one({"id": pid}, {"_id": 0})
    if not d:
        raise HTTPException(404, "Purchase payment not found")
    return PurchasePayment(**d)


@api_router.put("/purchase-payments/{pid}", response_model=PurchasePayment)
async def update_purchase_payment(pid: str, payload: PurchasePaymentBase):
    existing = await db.purchase_payments.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Purchase payment not found")
    old_pids = [a.get("purchase_id") for a in (existing.get("allocations") or []) if a.get("purchase_id")]
    p = PurchasePayment(**{**payload.model_dump(), "id": pid, "created_at": existing.get("created_at")}).model_dump()
    amt = float(p.get("amount") or 0)
    alloc = sum(float(a.get("amount") or 0) for a in (p.get("allocations") or []))
    p["allocated_total"] = round(alloc, 2)
    p["unallocated"] = round(max(0.0, amt - alloc), 2)
    await db.purchase_payments.update_one({"id": pid}, {"$set": p})
    new_pids = [a.get("purchase_id") for a in (p.get("allocations") or []) if a.get("purchase_id")]
    all_pids = list({*(old_pids or []), *(new_pids or [])})
    if all_pids:
        await _recompute_purchase_payment_aggregates(all_pids)
    return PurchasePayment(**p)


@api_router.delete("/purchase-payments/{pid}")
async def delete_purchase_payment(pid: str):
    existing = await db.purchase_payments.find_one({"id": pid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Purchase payment not found")
    pids = [a.get("purchase_id") for a in (existing.get("allocations") or []) if a.get("purchase_id")]
    await db.purchase_payments.delete_one({"id": pid})
    if pids:
        await _recompute_purchase_payment_aggregates(pids)
    return {"deleted": True}


# Register router AFTER all endpoints are defined


# ================================================================
# QUOTATIONS — native module (replaces embedded Samrat Glass ERP)
# ================================================================
class QuotationItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    product_name: str = ""
    description: Optional[str] = ""
    qty: float = 1
    rate: float = 0
    amount: float = 0


class Quotation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    quote_number: Optional[str] = ""  # auto-assigned on create if blank
    quote_date: Optional[str] = None
    valid_until: Optional[str] = None
    client_name: str = ""
    client_phone: Optional[str] = ""
    client_email: Optional[str] = ""
    billing_address: Optional[str] = ""
    billing_city: Optional[str] = ""
    billing_pincode: Optional[str] = ""
    shipping_same_as_billing: bool = True
    shipping_address: Optional[str] = ""
    shipping_city: Optional[str] = ""
    shipping_pincode: Optional[str] = ""
    items: List[QuotationItem] = []
    gst_rate: float = 18
    freight_type: Literal["extra", "included", "none"] = "extra"
    freight_amount: float = 0
    subtotal: float = 0
    tax_amount: float = 0
    total: float = 0
    notes: Optional[str] = ""
    terms: Optional[str] = ""
    status: Literal["Draft", "Sent", "Accepted", "Rejected", "Converted"] = "Draft"
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


def _compute_quotation_totals(q: dict) -> dict:
    """Recompute per-line amounts + subtotal + tax + total. Freight rules:
       - extra:    added on top of subtotal, taxed together with goods.
       - included: freight is already in the line rates — no separate row.
       - none:     no freight applied.
    """
    items = q.get("items") or []
    subtotal = 0.0
    for it in items:
        qty = float(it.get("qty") or 0)
        rate = float(it.get("rate") or 0)
        amt = round(qty * rate, 2)
        it["amount"] = amt
        subtotal += amt
    freight_type = q.get("freight_type") or "extra"
    freight = float(q.get("freight_amount") or 0) if freight_type == "extra" else 0.0
    taxable = subtotal + freight
    gst = float(q.get("gst_rate") or 0)
    tax = round(taxable * gst / 100.0, 2)
    total = round(taxable + tax, 2)
    q["subtotal"] = round(subtotal, 2)
    q["freight_amount"] = round(freight if freight_type == "extra" else float(q.get("freight_amount") or 0), 2)
    q["tax_amount"] = tax
    q["total"] = total
    return q


async def _next_quotation_number() -> str:
    """Generate the next SGE-YYYY-NNNN sequence based on quotations already stored."""
    year = datetime.now(timezone.utc).strftime("%Y")
    prefix = f"SGE-{year}-"
    # Find the highest existing sequence for this year
    latest = await db.quotations.find(
        {"quote_number": {"$regex": f"^{prefix}"}},
        {"quote_number": 1, "_id": 0},
    ).sort("quote_number", -1).limit(1).to_list(1)
    seq = 1
    if latest:
        try:
            seq = int(latest[0]["quote_number"].split("-")[-1]) + 1
        except Exception:
            seq = 1
    return f"{prefix}{seq:04d}"


@api_router.get("/quotations", response_model=List[Quotation])
async def list_quotations(
    search: Optional[str] = None,
    status: Optional[str] = None,
    limit: int = 500,
):
    query = {}
    if status and status != "all":
        query["status"] = status
    if search:
        query["$or"] = [
            {"client_name": {"$regex": search, "$options": "i"}},
            {"quote_number": {"$regex": search, "$options": "i"}},
        ]
    docs = await db.quotations.find(query, {"_id": 0}).sort("created_at", -1).limit(limit).to_list(limit)
    return docs


@api_router.get("/quotations/next-number")
async def next_quotation_number():
    return {"quote_number": await _next_quotation_number()}


@api_router.get("/quotations/{qid}", response_model=Quotation)
async def get_quotation(qid: str):
    doc = await db.quotations.find_one({"id": qid}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Quotation not found")
    return doc


@api_router.post("/quotations", response_model=Quotation)
async def create_quotation(payload: Quotation):
    now_iso = datetime.now(timezone.utc).isoformat()
    doc = payload.model_dump()
    if not doc.get("quote_number"):
        doc["quote_number"] = await _next_quotation_number()
    if not doc.get("quote_date"):
        doc["quote_date"] = now_iso
    doc["created_at"] = now_iso
    doc["updated_at"] = now_iso
    doc = _compute_quotation_totals(doc)
    await db.quotations.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.put("/quotations/{qid}", response_model=Quotation)
async def update_quotation(qid: str, payload: Quotation):
    existing = await db.quotations.find_one({"id": qid}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Quotation not found")
    doc = payload.model_dump()
    doc["id"] = qid
    doc["quote_number"] = doc.get("quote_number") or existing.get("quote_number")
    doc["created_at"] = existing.get("created_at")
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    doc = _compute_quotation_totals(doc)
    await db.quotations.replace_one({"id": qid}, doc)
    return doc


@api_router.delete("/quotations/{qid}")
async def delete_quotation(qid: str):
    r = await db.quotations.delete_one({"id": qid})
    if r.deleted_count == 0:
        raise HTTPException(404, "Quotation not found")
    return {"deleted": True}


@api_router.post("/quotations/{qid}/duplicate", response_model=Quotation)
async def duplicate_quotation(qid: str):
    src = await db.quotations.find_one({"id": qid}, {"_id": 0})
    if not src:
        raise HTTPException(404, "Quotation not found")
    now_iso = datetime.now(timezone.utc).isoformat()
    new_doc = {**src}
    new_doc["id"] = str(uuid.uuid4())
    new_doc["quote_number"] = await _next_quotation_number()
    new_doc["quote_date"] = now_iso
    new_doc["created_at"] = now_iso
    new_doc["updated_at"] = now_iso
    new_doc["status"] = "Draft"
    # Reset item ids so front-end doesn't collide
    for it in (new_doc.get("items") or []):
        it["id"] = str(uuid.uuid4())
    await db.quotations.insert_one(new_doc)
    new_doc.pop("_id", None)
    return new_doc


# --- Party Ledger v2 (unified) ---
app.include_router(make_party_ledger_v2_router(db), prefix="/api")


app.include_router(api_router)


async def _seed_payments_if_empty() -> int:
    """DISABLED — kept as a stub for reference.
    User policy: existing data must never be automatically recreated after deletion.
    Auto-restore removed. Use POST /api/seed?force=true&confirm_wipe=YES to explicitly
    reseed everything, or restore payments manually from backend/seed/pl_seed.json."""
    return 0


@app.on_event("startup")
async def _startup():
    order_count = await db.orders.count_documents({})
    tx_count = await db.transactions.count_documents({})

    if order_count == 0 and tx_count > 0:
        try:
            n = await _migrate_transactions_to_orders()
            logger.info(f"Auto-migrated {n} orders from legacy transactions.")
        except Exception as e:
            logger.error(f"Migration failed: {e}")
    elif order_count == 0 and tx_count == 0:
        try:
            r = await seed(force=False)
            logger.info(f"Auto-seed result: {r}")
        except Exception as e:
            logger.error(f"Auto-seed failed: {e}")

    # Migrate legacy order_payments → customer_payments (one-time, idempotent)
    try:
        n = await _migrate_order_payments_to_customer_payments()
        if n:
            logger.info(f"Migrated {n} legacy order payments to customer_payments collection.")
    except Exception as e:
        logger.error(f"Order-payment migration failed: {e}")

    # Backfill: existing orders without shipments should get one auto-shipment so
    # historical revenue reports don't drop to zero after this refactor.
    try:
        n = 0
        async for doc in db.orders.find({"$or": [
            {"shipments": {"$exists": False}},
            {"shipments": {"$size": 0}},
        ]}):
            doc.pop("_id", None)
            if not doc.get("items"):
                continue
            ship_items = [{"order_item_id": it["id"], "qty": float(it.get("qty") or 0)}
                          for it in doc["items"] if it.get("id")]
            auto = Shipment(
                date=doc.get("shipped_date") or doc.get("order_date"),
                items=[ShipmentItem(**s).model_dump() for s in ship_items],
                freight_paid=float(doc.get("freight_paid") or 0),
                freight_charged=float(doc.get("freight_charged") or 0),
                boxes_shipped=float(doc.get("boxes_shipped") or 0),
                transporter=doc.get("transporter") or "",
                lr_number=doc.get("lr_number") or "",
                remarks="Auto-created shipment (backfill during ERP refactor)",
            ).model_dump()
            doc["shipments"] = [auto]
            doc["status"] = "Fully Shipped"
            compute_order_aggregates(doc)
            await db.orders.update_one({"id": doc["id"]}, {"$set": doc})
            n += 1
        if n:
            logger.info(f"Backfilled auto-shipments on {n} legacy orders.")
    except Exception as e:
        logger.error(f"Shipment backfill failed: {e}")

    # Refresh aggregates & payment tallies
    try:
        n = await _refresh_stored_aggregates()
        if n:
            logger.info(f"Refreshed stored aggregates on {n} orders.")
        # Recompute payment aggregates for all orders (invoice may have changed)
        all_ids = await db.orders.distinct("id")
        await _recompute_payment_aggregates_for_orders(all_ids)
    except Exception as e:
        logger.error(f"Aggregate refresh failed: {e}")

    # Seed default accounts on fresh install
    try:
        n = await _seed_default_accounts_if_empty()
        if n:
            logger.info(f"Seeded {n} default accounts.")
    except Exception as e:
        logger.error(f"Account seed failed: {e}")

    # Party Ledger v2: seed system parties + derived parties from vendors/customers
    try:
        await ensure_party_ledger_bootstrap(db)
        n_parties = await db.parties.count_documents({})
        logger.info(f"Party Ledger v2 bootstrap OK — {n_parties} parties active.")
    except Exception as e:
        logger.error(f"Party Ledger v2 bootstrap failed: {e}")

    # Phase 2 (P1): Canonical party identity migration.
    try:
        report = await run_party_migration(db)
        logger.info(
            f"Phase 2 party migration OK — total parties: {report['parties_created']}, "
            f"vendors_linked: {report['vendors_linked']}, "
            f"customers_linked: {report['customers_linked']}, "
            f"FF aliases resolved: {report['ff_aliases_resolved']}, "
            f"exact duplicates merged: {report['exact_duplicates_merged']}, "
            f"probable duplicates flagged: {report['probable_duplicates_flagged']}, "
            f"unmatched legacy names: {len(report['unmatched_legacy_names'])}"
        )
        await db.admin_migration_reports.insert_one({
            "id": str(uuid.uuid4()),
            "phase": "P1_party_identity",
            "created_at": datetime.now(timezone.utc).isoformat(),
            **report,
        })
    except Exception as e:
        logger.error(f"Phase 2 party migration failed: {e}")

    # Compound unique index on (type, normalized_name) — enforces race-safe
    # canonical resolution for all subsequent writes. Created AFTER migration
    # so any exact duplicates have already been merged. Partial filter
    # excludes rows without a normalized_name (legacy compat).
    try:
        await db.parties.create_index(
            [("type", 1), ("normalized_name", 1)],
            unique=True,
            name="party_identity_uidx",
            partialFilterExpression={
                "normalized_name": {"$exists": True, "$type": "string"},
            },
        )
    except Exception as e:
        logger.warning(f"Phase 2 party unique index setup skipped: {e}")

    # Phase 3 (P1): Transfers indexes + one-time migration from legacy
    # cash_book_entries[kind='transfer'] into db.transfers. Deterministic on
    # `legacy_cbe_id` so legacy and migrated versions never appear together.
    try:
        await ensure_transfer_indexes(db)
        t_report = await run_transfer_migration(db)
        logger.info(
            f"Phase 3 transfer migration OK — created: {t_report['created']}, "
            f"already_migrated: {t_report['already_migrated']}, "
            f"skipped_no_account: {t_report['skipped_no_account']}"
        )
        await db.admin_migration_reports.insert_one({
            "id": str(uuid.uuid4()),
            "created_at": datetime.now(timezone.utc).isoformat(),
            **t_report,
        })
    except Exception as e:
        logger.error(f"Phase 3 transfer migration failed: {e}")

    # P0: stamp every pre-existing db.payments row with source='legacy_migrated'
    # so it can be surfaced in the Cash Book timeline as read-only, without
    # ever entering canonical KPI computations. Idempotent.
    try:
        res = await db.payments.update_many(
            {"source": {"$exists": False}},
            {"$set": {"source": "legacy_migrated"}},
        )
        if res.modified_count:
            logger.info(
                f"P0: stamped {res.modified_count} legacy Cash Book rows as legacy_migrated."
            )
    except Exception as e:
        logger.error(f"P0 legacy stamp failed: {e}")


@app.on_event("shutdown")
async def _shutdown():
    client.close()
